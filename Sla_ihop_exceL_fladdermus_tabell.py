import os
import re
import sys
import subprocess
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ================== UI ==================
TITLE_OPEN = "Välj en eller flera Excel-filer (.xlsx, samma struktur)"
TITLE_SAVE = "Välj var du vill spara sammanställningen"
DEFAULT_OUT = "sammanstallning_fladdermus.xlsx"

# ================== Kolory / style ==================
COLOR_FORBI = "FFE7E6E6"     # Förbiflygande (jasnoszary)
COLOR_SOC   = "FFFF0000"     # Socialt (czerwony)
COLOR_FODO  = "FFFFC000"     # Fodosökande (żółty)
COLOR_HEADER_BG = "FF595959" # nagłówek tło
COLOR_HEADER_FG = "FFFFFFFF" # nagłówek font

FILL_FORBI = PatternFill("solid", fgColor=COLOR_FORBI)
FILL_SOC   = PatternFill("solid", fgColor=COLOR_SOC)
FILL_FODO  = PatternFill("solid", fgColor=COLOR_FODO)
FILL_HDR   = PatternFill("solid", fgColor=COLOR_HEADER_BG)

BORDER_MEDIUM = Side(style="medium", color="FF000000")

# ================== Słownik: Łacina → Szwedzki ==================
LATIN_TO_SV = {
    "Barbastella barbastellus": "barbastell",
    "Eptesicus nilssonii": "nordfladdermus",
    "Eptesicus serotinus": "sydfladdermus",
    "Myotis alcathoe": "nymffladdermus",
    "Myotis bechsteinii": "Bechsteins fladdermus",
    "Myotis brandtii": "tajgafladdermus",
    "Myotis dasycneme": "dammfladdermus",
    "Myotis daubentonii": "vattenfladdermus",
    "Myotis myotis": "större musöra",
    "Myotis mystacinus": "mustaschfladdermus",
    "Myotis nattereri": "fransfladdermus",
    "Myotis mystacinus/brandtii": "mustasch/tajgafladdermus",
    "Nyctalus leisleri": "mindre brunfladdermus",
    "Nyctalus noctula": "större brunfladdermus",
    "Pipistrellus kuhlii": "parkpipistrell",
    "Pipistrellus nathusii": "trollpipistrell",
    "Pipistrellus pipistrellus": "sydpipistrell",
    "Pipistrellus pygmaeus": "dvärgpipistrell",
    "Plecotus auritus": "brunlångöra",
    "Plecotus austriacus": "grålångöra",
    "Vespertilio murinus": "gråskimlig fladdermus",
    "Nyctaloid": None,
    "Chiroptera": None,
}
SPECIAL_TAIL = {"nyctaloid", "chiroptera"}  # na koniec

def display_label_multiline(latin: str) -> str:
    """Zwraca 'Szwedzka,\nŁacińska' lub samo 'Łacińska' (dla Nyctaloid/Chiroptera)."""
    latin = latin.strip()
    sv = LATIN_TO_SV.get(latin, None)
    if sv:
        return f"{sv.capitalize()},\n{latin}"
    return latin

# ================== Pomocnicze ==================
def safe_sheet_name(path: str, used: set) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    base = re.sub(r'[:\\\/\?\*\[\]]', '_', base).strip()[:31] or "Ark"
    cand = base; i = 2
    while cand in used or not cand:
        suf = f"_{i}"
        cand = (base[: (31 - len(suf))] + suf) if len(base) + len(suf) > 31 else (base + suf)
        i += 1
    used.add(cand); return cand

def extract_species_and_type(manual_id_value):
    """Rozbija MANUAL ID. Mapowanie typów: Förbiflygande/Socialt/Fodosökande."""
    if pd.isna(manual_id_value): return []
    out = []
    for raw in str(manual_id_value).split(","):
        entry = raw.strip()
        if not entry: continue
        upper = entry.upper()
        if "FOD" in upper: typ = "Fodosökande"
        elif "SOC" in upper: typ = "Socialt"
        else: typ = "Förbiflygande"
        species = re.sub(r'\b(FOD|SOC)\b', "", entry, flags=re.IGNORECASE).strip(" ,;")
        out.append((species, typ))
    return out

def open_file(path):
    try:
        if sys.platform.startswith("win"): os.startfile(path)
        elif sys.platform == "darwin": subprocess.run(["open", path])
        else: subprocess.run(["xdg-open", path])
    except Exception as e:
        print(f"Kan inte öppna filen automatiskt: {e}")

# ================== Wybór plików (.xlsx tylko) ==================
root = tk.Tk(); root.withdraw()
input_files = filedialog.askopenfilenames(
    title=TITLE_OPEN, filetypes=[("Excel-filer (.xlsx)", "*.xlsx")]
)
if not input_files: print("Ingen fil vald. Avslutar."); raise SystemExit
input_files = [p for p in input_files if p.lower().endswith(".xlsx")]
if not input_files: print("Endast .xlsx stöds. Avslutar."); raise SystemExit

out_path = filedialog.asksaveasfilename(
    title=TITLE_SAVE, defaultextension=".xlsx",
    initialfile=DEFAULT_OUT, filetypes=[("Excel-fil (.xlsx)", "*.xlsx")]
)
if not out_path: print("Ingen plats vald. Avslutar."); raise SystemExit

# ================== Przetwarzanie ==================
used_sheet_names = set()
sheets_to_write = []       # [(sheet_name, df)]
counts_per_file = {}       # {sheet_name: {(latin, type): count}}
total_ljud_per_file = {}   # {sheet_name: total_rows}
all_species_latin = set()

for path in input_files:
    sheet_name = safe_sheet_name(path, used_sheet_names)
    df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    sheets_to_write.append((sheet_name, df))
    total_ljud_per_file[sheet_name] = int(len(df))  # z NOISE

    if "MANUAL ID" not in df.columns:
        counts_per_file[sheet_name] = {}
        continue

    df["__list"] = df["MANUAL ID"].map(extract_species_and_type)
    long = df.explode("__list")
    long = long[long["__list"].notna()]
    if long.empty:
        counts_per_file[sheet_name] = {}
        continue

    # kolumny: łacina + Beteendetyper
    long[["ArtLatin", "Beteendetyper"]] = pd.DataFrame(long["__list"].tolist(), index=long.index)

    # wyklucz 'Noise'
    mask_ok = long["ArtLatin"].astype(str).str.strip().str.lower() != "noise"
    long = long[mask_ok]

    all_species_latin.update(long["ArtLatin"].astype(str).str.strip().tolist())

    grp = long.groupby(["ArtLatin", "Beteendetyper"]).size()
    counts_per_file[sheet_name] = {(sp, typ): int(n) for (sp, typ), n in grp.items()}

# ================== Budowa DataFrame „Översikt” ==================
type_order = ["Förbiflygande", "Socialt", "Fodosökande"]

def species_sort_key(latin: str):
    return (latin.strip().lower() in SPECIAL_TAIL, latin.casefold())

species_sorted_latin = sorted(all_species_latin, key=species_sort_key)
file_cols = list(counts_per_file.keys())

rows_data = []
for latin in species_sorted_latin:
    disp = display_label_multiline(latin)  # << dwie linie w komórce
    for typ in type_order:
        row = {"Art": disp, "Beteendetyper": typ}
        for col in file_cols:
            val = counts_per_file.get(col, {}).get((latin, typ), 0)
            row[col] = ("" if val == 0 else int(val))  # pusto zamiast 0
        rows_data.append(row)

# sumy (bez Noise)
sum_row = {"Art": "", "Beteendetyper": "Fladdermusregistreringar"}
for col in file_cols:
    sum_row[col] = int(sum(counts_per_file.get(col, {}).values()))
rows_data.append(sum_row)

# total nagrań (z Noise)
tot_row = {"Art": "", "Beteendetyper": "Total antal ljud"}
for col in file_cols:
    tot_row[col] = int(total_ljud_per_file.get(col, 0))
rows_data.append(tot_row)

overview_df = pd.DataFrame(rows_data, columns=["Art", "Beteendetyper"] + file_cols)

# ================== Zapis wstępny (pandas -> .xlsx) ==================
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    overview_df.to_excel(writer, sheet_name="Översikt", index=False)
    for sheet_name, df_orig in sheets_to_write:
        df_orig.to_excel(writer, sheet_name=sheet_name, index=False)

# ================== Stylowanie i grube linie (openpyxl) ==================
wb = load_workbook(out_path)
ws = wb["Översikt"]

max_row = ws.max_row
max_col = ws.max_column
header_row = 1
data_start = header_row + 1

num_species_rows = len(species_sorted_latin) * 3
sum_row_idx   = data_start + num_species_rows          # "Fladdermusregistreringar"
total_row_idx = sum_row_idx + 1                        # "Total antal ljud"

# Nagłówek
for c in range(1, max_col + 1):
    cell = ws.cell(row=header_row, column=c)
    cell.fill = FILL_HDR
    cell.font = Font(color=COLOR_HEADER_FG, bold=True)
    cell.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[header_row].height = 18

# Szerokości kolumn
ws.column_dimensions["A"].width = 44    # Art (dwuliniowe)
ws.column_dimensions["B"].width = 20    # Beteendetyper
# Auto-dopasuj kolumny z plikami (nagłówki C..)
for idx, col_name in enumerate(file_cols, start=3):
    header_text = str(col_name)
    width = max(12, min(50, int(len(header_text) * 1.1)))
    ws.column_dimensions[get_column_letter(idx)].width = width

# Merge + wyśrodkowanie pionowe + zawijanie tekstu w kolumnie A (Art)
if num_species_rows > 0:
    current = data_start
    merge_end_limit = sum_row_idx - 1
    while current <= merge_end_limit:
        art_val = ws[f"A{current}"].value
        if not art_val:
            current += 1
            continue
        end = current
        while end + 1 <= merge_end_limit and ws[f"A{end+1}"].value == art_val:
            end += 1
        if end > current:
            ws.merge_cells(start_row=current, start_column=1, end_row=end, end_column=1)
        ws.cell(row=current, column=1).alignment = Alignment(vertical="center", wrap_text=True)
        current = end + 1

# Kolorowanie wg kategorii (kol. B = Beteendetyper)
def fill_for_type(t):
    if t == "Förbiflygande": return FILL_FORBI
    if t == "Socialt":       return FILL_SOC
    if t == "Fodosökande":   return FILL_FODO
    return None

for r in range(data_start, sum_row_idx):
    typ = ws.cell(row=r, column=2).value
    fill = fill_for_type(typ)
    if fill:
        ws.cell(row=r, column=2).fill = fill
        for c in range(3, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val not in (None, "", 0):
                ws.cell(row=r, column=c).fill = fill

# Wiersze sum – pogrubienie
for r in (sum_row_idx, total_row_idx):
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

# ====== GRUBE LINIE ======

# 1) Poziome nad każdym blokiem gatunku
for i in range(len(species_sorted_latin)):
    top_row = data_start + i * 3
    for c in range(1, max_col + 1):
        old = ws.cell(row=top_row, column=c).border
        ws.cell(row=top_row, column=c).border = Border(
            left=old.left, right=old.right, top=BORDER_MEDIUM, bottom=old.bottom
        )

# 2) Gruba nad „Fladdermusregistreringar” i gruba na dole pod „Total antal ljud”
for c in range(1, max_col + 1):
    old = ws.cell(row=sum_row_idx, column=c).border
    ws.cell(row=sum_row_idx, column=c).border = Border(
        left=old.left, right=old.right, top=BORDER_MEDIUM, bottom=old.bottom
    )
    old = ws.cell(row=total_row_idx, column=c).border
    ws.cell(row=total_row_idx, column=c).border = Border(
        left=old.left, right=old.right, top=old.top, bottom=BORDER_MEDIUM
    )

# 3) Pionowe grube linie:
#    - między kolumnami plików (C..),
#    - gruba lewa krawędź kol. A (zewnętrzna),
#    - **gruba linia po kolumnie A** (lewy brzeg kol. B),
#    - gruba prawa krawędź tabeli (ostatnia kolumna).
for c in range(3, max_col + 1):
    for r in range(header_row, max_row + 1):
        old = ws.cell(row=r, column=c).border
        ws.cell(row=r, column=c).border = Border(
            left=BORDER_MEDIUM, right=old.right, top=old.top, bottom=old.bottom
        )
for r in range(header_row, max_row + 1):
    # lewy brzeg tabeli (A)
    oldA = ws.cell(row=r, column=1).border
    ws.cell(row=r, column=1).border = Border(
        left=BORDER_MEDIUM, right=oldA.right, top=oldA.top, bottom=oldA.bottom
    )
    # gruba linia po kolumnie A → lewy brzeg kolumny B
    oldB = ws.cell(row=r, column=2).border
    ws.cell(row=r, column=2).border = Border(
        left=BORDER_MEDIUM, right=oldB.right, top=oldB.top, bottom=oldB.bottom
    )
    # prawa krawędź tabeli
    oldR = ws.cell(row=r, column=max_col).border
    ws.cell(row=r, column=max_col).border = Border(
        left=oldR.left, right=BORDER_MEDIUM, top=oldR.top, bottom=oldR.bottom
    )

wb.save(out_path)
wb.close()

# ================== Otwórz plik automatycznie ==================
open_file(out_path)
print(f"Klar! Sparad fil: {out_path}")
