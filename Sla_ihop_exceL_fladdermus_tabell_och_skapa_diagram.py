import os
import re
import sys
import math
import subprocess
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from matplotlib.ticker import MaxNLocator

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk, colorchooser

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ================== Gränssnitt (UI) ==================
TITLE_OPEN = "Välj en eller flera Excel-filer (.xlsx, samma struktur)"
TITLE_SAVE = "Välj var du vill spara sammanställningen (ange basnamn)"
DEFAULT_OUT = "sammanstallning_fladdermus.xlsx"

# ================== Färger – enhetligt HEX-format ==================
# Vi använder HEX #RRGGBB överallt; till Excel konverterar vi till ARGB ("FF"+RRGGBB).

# Rubrikrad (Översikt)
HEX_HEADER_BG = "#595959"
HEX_HEADER_FG = "#FFFFFF"

# Tabell: gemensam färg för Förbiflygande (samma i NVI och ART)
HEX_TABLE_FORBI = "#E7E6E6"

# Palett NVI (stapeldiagram)
HEX_NVI_SOC   = "#FF0000"  # Socialt
HEX_NVI_FODO  = "#FFC000"  # Födosökande
HEX_NVI_FORBI = "#A9A9A9"  # Förbiflygande (i diagram; i tabell används HEX_TABLE_FORBI)

# Palett ART (stapeldiagram)
HEX_ART_SOC   = "#EB09D8"  # Socialt
HEX_ART_FODO  = "#D98FD3"  # Födosökande
HEX_ART_FORBI = "#ABAAA9"  # Förbiflygande (i diagram; i tabell används HEX_TABLE_FORBI)

# Kantlinjer i Excel
BORDER_MEDIUM = Side(style="medium", color="FF000000")

def hex_to_argb(hex_rgb: str) -> str:
    """Konverterar '#RRGGBB' → 'FFRRGGBB' (ARGB) för Excel-färger."""
    h = hex_rgb.strip().lstrip("#")
    if len(h) != 6:
        raise ValueError(f"Ogiltig HEX: {hex_rgb}")
    return "FF" + h.upper()

def fill_from_hex(hex_rgb: str) -> PatternFill:
    """Skapar PatternFill från HEX (#RRGGBB)."""
    return PatternFill("solid", fgColor=hex_to_argb(hex_rgb))

FILL_HDR = fill_from_hex(HEX_HEADER_BG)

# ================== Ordbok: Latin → Svenska ==================
LATIN_TO_SV = {
    "Barbastella barbastellus": "Barbastell",
    "Eptesicus nilssonii": "Nordfladdermus",
    "Eptesicus serotinus": "Sydfladdermus",
    "Myotis alcathoe": "Nymffladdermus",
    "Myotis bechsteinii": "Bechsteins fladdermus",
    "Myotis brandtii": "Tajgafladdermus",
    "Myotis dasycneme": "Dammfladdermus",
    "Myotis daubentonii": "Vattenfladdermus",
    "Myotis myotis": "Större musöra",
    "Myotis mystacinus": "Mustaschfladdermus",
    "Myotis nattereri": "Fransfladdermus",
    "Myotis mystacinus/brandtii": "Mustasch/Tajgafladdermus",
    "Nyctalus leisleri": "Mindre brunfladdermus",
    "Nyctalus noctula": "Större brunfladdermus",
    "Pipistrellus kuhlii": "Parkpipistrell",
    "Pipistrellus nathusii": "Trollpipistrell",
    "Pipistrellus pipistrellus": "Sydpipistrell",
    "Pipistrellus pygmaeus": "Dvärgpipistrell",
    "Plecotus auritus": "Brunlångöra",
    "Plecotus austriacus": "Grålångöra",
    "Vespertilio murinus": "Gråskimlig fladdermus",
    "Nyctaloid": None,
    "Chiroptera": None,
}
SPECIAL_TAIL = {"nyctaloid", "chiroptera"}  # sorteras sist i artlistor

# ================== Hjälpfunktioner (gemensamma) ==================
def safe_sheet_name(path: str, used: set) -> str:
    """Gör kalkylbladsnamn Excel-kompatibelt (max 31 tecken, förbjudna tecken ersätts)."""
    base = os.path.splitext(os.path.basename(path))[0]
    base = re.sub(r'[:\\\/\?\*\[\]]', '_', base).strip()[:31] or "Ark"
    cand = base; i = 2
    while cand in used or not cand:
        suf = f"_{i}"
        cand = (base[: (31 - len(suf))] + suf) if len(base) + len(suf) > 31 else (base + suf)
        i += 1
    used.add(cand); return cand

def display_label_multiline(latin: str) -> str:
    """Returnerar 'Svenskt namn,\\nLatinskt namn' (om känt), annars originaltexten."""
    latin = str(latin).strip()
    sv = LATIN_TO_SV.get(latin, None)
    if sv:
        return f"{sv.capitalize()},\n{latin}"
    return latin

def format_title(species_latin: str, total_count: int) -> str:
    """Titel för artspecifika diagram: '<svenskt> (<latinskt>), antal observerade beteenden: NN'."""
    sv = LATIN_TO_SV.get(species_latin)
    if sv:
        return f"{sv} ({species_latin}), antal observerade beteenden: {int(total_count)}"
    return f"{species_latin}, antal observerade beteenden: {int(total_count)}"

def extract_species_and_type(manual_id_value):
    """Tolkar fältet MANUAL ID till lista av (art, beteendetyp: Förbiflygande/Socialt/Födosökande)."""
    if pd.isna(manual_id_value): return []
    out = []
    for raw in str(manual_id_value).split(","):
        entry = raw.strip()
        if not entry: continue
        upper = entry.upper()
        if "FOD" in upper: typ = "Födosökande"
        elif "SOC" in upper: typ = "Socialt"
        else: typ = "Förbiflygande"
        species = re.sub(r'\b(FOD|SOC)\b', "", entry, flags=re.IGNORECASE).strip(" ,;")
        out.append((species, typ))
    return out

def open_file(path):
    """Försöker öppna fil i OS:et för snabb visuell kontroll."""
    try:
        if sys.platform.startswith("win"): os.startfile(path)
        elif sys.platform == "darwin": subprocess.run(["open", path])
        else: subprocess.run(["xdg-open", path])
    except Exception as e:
        print(f"Kan inte öppna filen automatiskt: {e}")

def species_sort_key(latin: str):
    """Sorteringsnyckel: A–Ö, men Nyctaloid/Chiroptera alltid sist."""
    return (str(latin).strip().lower() in SPECIAL_TAIL, str(latin).casefold())

def safe_filename(s):
    """Gör sträng filnamnssäker (tar bort/ersätter otillåtna tecken)."""
    return re.sub(r'[\\/:\*\?"<>\|]', '_', str(s))

# --- Tidsparsing (sträng, datetime, Excel-float) ---
def _hm_from_any(val):
    """Returnerar (timme, minut) eller None om värdet inte kan tolkas som tid."""
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)) or (isinstance(val, str) and val.strip() == ""):
            return None
        if hasattr(val, "hour") and hasattr(val, "minute"):
            return int(val.hour), int(val.minute)
        if isinstance(val, (int, float)) and not pd.isna(val):
            frac = float(val) % 1.0
            secs = int(round(frac * 24 * 60 * 60))
            h = (secs // 3600) % 24
            m = (secs % 3600) // 60
            return int(h), int(m)
        s = str(val).strip()
        t = pd.to_datetime(s, format="%H:%M:%S", errors="coerce")
        if pd.isna(t):
            t = pd.to_datetime(s, format="%H:%M", errors="coerce")
        if pd.isna(t):
            return None
        return int(t.hour), int(t.minute)
    except Exception:
        return None

def str_to_dt(time_val):
    """Bygger ett datetime (med konstgjort datum) från valfri tolkbar tid."""
    hm = _hm_from_any(time_val)
    if hm is None:
        return None
    h, m = hm
    fake_date = "2000-01-02" if h < 12 else "2000-01-01"
    return datetime.strptime(f"{fake_date} {h:02d}:{m:02d}", "%Y-%m-%d %H:%M")

def round_down_15(dt):
    """Rundar ned till närmaste 15-minutersintervall."""
    return dt.replace(minute=(dt.minute // 15) * 15, second=0, microsecond=0)

def round_up_15(dt):
    """Rundar upp till närmaste 15-minutersintervall."""
    if dt.minute % 15 != 0 or dt.second > 0 or dt.microsecond > 0:
        dt = dt + timedelta(minutes=15 - (dt.minute % 15), seconds=-dt.second, microseconds=-dt.microsecond)
    return dt.replace(second=0, microsecond=0)

def interval_to_sortkey(interval):
    """Konverterar 'HH:MM' till datetime (med konstgjort datum) för sortering/jämförelse."""
    try:
        t = pd.to_datetime(str(interval), format="%H:%M", errors="coerce")
        if pd.isna(t):
            return None
        h, m = int(t.hour), int(t.minute)
        fake_date = "2000-01-02" if h < 12 else "2000-01-01"
        return datetime.strptime(f"{fake_date} {h:02d}:{m:02d}", "%Y-%m-%d %H:%M")
    except Exception:
        return None

def detect_column(df, candidates):
    """Hittar första kolumn vars namn (case-insensitivt) matchar en av kandidaterna."""
    lowmap = {str(c).strip().lower(): c for c in df.columns}
    for k in candidates:
        if k in lowmap:
            return lowmap[k]
    return None

def count_nights(df):
    """
    Räknar antal unika fältnätter i en fil.
    - Använder DATE/Datum och TIME/Tid (tider < 12 → räknas till föregående natt).
    - Om datum saknas returneras None.
    """
    date_col = detect_column(df, ["date", "datum"])
    if not date_col:
        return None
    time_col = detect_column(df, ["time", "tid"])

    dates = pd.to_datetime(df[date_col], errors="coerce")
    if time_col:
        hm = df[time_col].apply(_hm_from_any)
        hours = hm.apply(lambda x: x[0] if isinstance(x, tuple) else None)
        shift = hours.apply(lambda h: (h is not None) and (h < 12))
        night_key = (dates.dt.normalize() - pd.to_timedelta(shift.fillna(False).astype(int), unit="D")).dt.date
    else:
        night_key = dates.dt.normalize().dt.date

    nights = pd.Series(night_key).dropna().nunique()
    return int(nights) if nights else None

# ================== GUI: samlingspanel för input, utdata, tider och färger ==================
def _validate_hex(s):
    s = (s or "").strip()
    if not s:
        return None
    if s.startswith("#"):
        s = s[1:]
    if len(s) != 6 or any(c not in "0123456789abcdefABCDEF" for c in s):
        return None
    return "#" + s.upper()

def gui_collect_settings(
    default_colors_nvi=("FF0000","FFC000","A9A9A9"),   # Socialt, Födosökande, Förbiflygande
    default_colors_art=("EB09D8","D98FD3","ABAAA9"),   # Socialt, Födosökande, Förbiflygande
    default_basename="sammanstallning_fladdermus"
):
    """Visar ett fönster och returnerar en dict med valen. Avbryt → sys.exit(0)."""
    root = tk.Tk()
    root.title("Fladdermus – sammanställning & diagram (GUI)")
    root.geometry("860x760")
    root.minsize(800, 640)

    main = ttk.Frame(root, padding=12)
    main.pack(fill="both", expand=True)

    # ---------- INDATA ----------
    lf_in = ttk.LabelFrame(main, text="Indatafiler")
    lf_in.pack(fill="both", expand=False, padx=0, pady=(0,10))

    files_list = tk.Listbox(lf_in, height=7, selectmode="extended")
    files_list.grid(row=0, column=0, rowspan=3, sticky="nsew", padx=(8,8), pady=8)
    lf_in.columnconfigure(0, weight=1)
    lf_in.rowconfigure(0, weight=1)

    def add_files():
        paths = filedialog.askopenfilenames(
            title="Välj en eller flera Excel-filer",
            filetypes=[("Excel-filer", "*.xlsx"), ("Alla filer", "*.*")]
        )
        if not paths: return
        for p in paths:
            files_list.insert(tk.END, p)

    def clear_files():
        files_list.delete(0, tk.END)

    ttk.Button(lf_in, text="Lägg till filer…", command=add_files).grid(row=0, column=1, sticky="ew", padx=(0,8), pady=(8,4))
    ttk.Button(lf_in, text="Rensa listan", command=clear_files).grid(row=1, column=1, sticky="ew", padx=(0,8))

    # ---------- UTDATA ----------
    lf_out = ttk.LabelFrame(main, text="Utdatakatalog och basnamn")
    lf_out.pack(fill="x", expand=False, pady=(0,10))

    var_outdir = tk.StringVar(value="")
    var_basename = tk.StringVar(value=default_basename)

    def pick_outdir():
        d = filedialog.askdirectory(title="Välj mapp där 'Results' ska skapas")
        if d:
            var_outdir.set(d)

    ttk.Label(lf_out, text="Basnamn (utan ändelse):").grid(row=0, column=0, sticky="w", padx=8, pady=(8,4))
    ttk.Entry(lf_out, textvariable=var_basename).grid(row=0, column=1, sticky="ew", padx=8, pady=(8,4))
    ttk.Label(lf_out, text="Mapp för 'Results/':").grid(row=1, column=0, sticky="w", padx=8)
    ttk.Entry(lf_out, textvariable=var_outdir).grid(row=1, column=1, sticky="ew", padx=8)
    ttk.Button(lf_out, text="Välj mapp…", command=pick_outdir).grid(row=1, column=2, sticky="ew", padx=(0,8))
    lf_out.columnconfigure(1, weight=1)

    # ---------- DIAGRAM ----------
    lf_plot = ttk.LabelFrame(main, text="Diagram (valfritt)")
    lf_plot.pack(fill="x", expand=False, pady=(0,10))

    var_do_plots = tk.BooleanVar(value=True)
    var_time_mode = tk.StringVar(value="auto")  # "auto" eller "manual"
    var_tstart = tk.StringVar(value="22:00")
    var_tend   = tk.StringVar(value="02:00")

    def toggle_time_entries():
        state = "disabled" if var_time_mode.get()=="auto" or not var_do_plots.get() else "normal"
        ent_start.config(state=state)
        ent_end.config(state=state)

    ttk.Checkbutton(lf_plot, text="Generera diagram", variable=var_do_plots, command=toggle_time_entries)\
        .grid(row=0, column=0, sticky="w", padx=8, pady=(8,4), columnspan=4)

    ttk.Radiobutton(lf_plot, text="X-axel: automatisk (från data)", value="auto", variable=var_time_mode, command=toggle_time_entries)\
        .grid(row=1, column=0, sticky="w", padx=8)
    ttk.Radiobutton(lf_plot, text="X-axel: eget intervall", value="manual", variable=var_time_mode, command=toggle_time_entries)\
        .grid(row=1, column=1, sticky="w")

    ttk.Label(lf_plot, text="Starttid (HH:MM):").grid(row=2, column=0, sticky="e", padx=8, pady=(4,8))
    ent_start = ttk.Entry(lf_plot, textvariable=var_tstart, width=10)
    ent_start.grid(row=2, column=1, sticky="w", pady=(4,8))

    ttk.Label(lf_plot, text="Sluttid (HH:MM):").grid(row=2, column=2, sticky="e", padx=8, pady=(4,8))
    ent_end = ttk.Entry(lf_plot, textvariable=var_tend, width=10)
    ent_end.grid(row=2, column=3, sticky="w", pady=(4,8))

    # Valfri målmapp för diagram (default: Results/diagramer)
    ttk.Label(lf_plot, text="Bas-mapp för diagram (valfritt):").grid(row=3, column=0, sticky="e", padx=8, pady=(0,8))
    var_diagdir = tk.StringVar(value="")
    ent_diag = ttk.Entry(lf_plot, textvariable=var_diagdir)
    ent_diag.grid(row=3, column=1, sticky="ew", padx=8, pady=(0,8), columnspan=2)
    def pick_diagdir():
        d = filedialog.askdirectory(title="Välj bas-mapp för diagram (om tomt → Results/diagramer)")
        if d:
            var_diagdir.set(d)
    ttk.Button(lf_plot, text="Välj mapp…", command=pick_diagdir).grid(row=3, column=3, sticky="ew", pady=(0,8))
    lf_plot.columnconfigure(1, weight=1)

    # ---------- FÄRGER (z paletą) ----------
    lf_colors = ttk.LabelFrame(main, text="Färger (valfritt – lämna tomt för standard)")
    lf_colors.pack(fill="x", expand=False)

    # Pomocnicze: aktualizacja podglądu przy zmianie wartości
    def bind_preview(var, preview_widget):
        def _cb(*_):
            hx = _validate_hex(var.get())
            preview_widget.config(background=(hx or "#FFFFFF"))
        var.trace_add("write", _cb)
        _cb()

    def color_row(row, label_text, vars_tuple):
        ttk.Label(lf_colors, text=label_text).grid(row=row, column=0, sticky="w", padx=8, pady=(8,2))
        entries = []
        previews = []
        def mk_one(col_ix, var):
            # Entry
            e = ttk.Entry(lf_colors, textvariable=var, width=10)
            e.grid(row=row, column=1 + col_ix*3, padx=(4,2), pady=(8,2), sticky="w")
            entries.append(e)
            # Preview (Label jako próbnik koloru)
            prev = tk.Label(lf_colors, text="  ", width=3, relief="groove")
            prev.grid(row=row, column=2 + col_ix*3, padx=(0,2), pady=(8,2), sticky="w")
            previews.append(prev)
            bind_preview(var, prev)
            # Button „Välj…”
            def choose():
                init = _validate_hex(var.get()) or "#FFFFFF"
                rgb, hx = colorchooser.askcolor(color=init, title="Välj färg")
                if hx:
                    var.set(hx.upper())
            ttk.Button(lf_colors, text="Välj…", command=choose)\
                .grid(row=row, column=3 + col_ix*3, padx=(0,6), pady=(8,2), sticky="w")

        for i, v in enumerate(vars_tuple):
            mk_one(i, v)

    # Zmienne HEX (startowe: standard)
    var_nvi_soc   = tk.StringVar(value="#FF0000")
    var_nvi_fodo  = tk.StringVar(value="#FFC000")
    var_nvi_forbi = tk.StringVar(value="#A9A9A9")
    var_art_soc   = tk.StringVar(value="#EB09D8")
    var_art_fodo  = tk.StringVar(value="#D98FD3")
    var_art_forbi = tk.StringVar(value="#ABAAA9")

    color_row(0, "NVI – Socialt / Födosökande / Förbiflygande (#RRGGBB):",
              (var_nvi_soc, var_nvi_fodo, var_nvi_forbi))
    color_row(1, "ART – Socialt / Födosökande / Förbiflygande (#RRGGBB):",
              (var_art_soc, var_art_fodo, var_art_forbi))

    # Rozciąganie kolumn
    for c in range(10):
        lf_colors.columnconfigure(c, weight=0)
    lf_colors.columnconfigure(9, weight=1)

    # ---------- KNAPPAR ----------
    btns = ttk.Frame(main); btns.pack(fill="x", pady=(12,0))
    ok_pressed = {"done": False}
    settings = {}

    def on_cancel():
        root.destroy()
        sys.exit(0)

    def on_ok():
        files = list(files_list.get(0, tk.END))
        if not files:
            messagebox.showwarning("GUI", "Välj minst en indatafil.")
            return

        outdir = var_outdir.get().strip()
        if not outdir:
            messagebox.showwarning("GUI", "Välj mapp där 'Results/' ska skapas.")
            return

        basename = var_basename.get().strip() or "sammanstallning_fladdermus"

        # Tider
        do_plots = bool(var_do_plots.get())
        tmode = var_time_mode.get()
        tstart = var_tstart.get().strip()
        tend   = var_tend.get().strip()
        custom_range = None
        if do_plots and tmode == "manual":
            try:
                pd.to_datetime(tstart, format="%H:%M")
                pd.to_datetime(tend,   format="%H:%M")
            except Exception:
                messagebox.showerror("GUI", "Felaktigt tidsformat. Använd HH:MM (t.ex. 22:15).")
                return
            custom_range = (tstart, tend)

        # Färger – tom = standard (None)
        nvi_soc   = _validate_hex(var_nvi_soc.get())
        nvi_fodo  = _validate_hex(var_nvi_fodo.get())
        nvi_forbi = _validate_hex(var_nvi_forbi.get())
        art_soc   = _validate_hex(var_art_soc.get())
        art_fodo  = _validate_hex(var_art_fodo.get())
        art_forbi = _validate_hex(var_art_forbi.get())

        diag_base = var_diagdir.get().strip() or None

        settings.update({
            "input_files": files,
            "base_dir": outdir,
            "base_name": basename,
            "do_plots": do_plots,
            "custom_time_range": custom_range,   # None = auto
            "diagram_base": diag_base,
            "colors": {
                "NVI": {"Socialt": nvi_soc, "Födosökande": nvi_fodo, "Förbiflygande": nvi_forbi},
                "ART": {"Socialt": art_soc, "Födosökande": art_fodo, "Förbiflygande": art_forbi},
            }
        })
        ok_pressed["done"] = True
        root.destroy()

    ttk.Button(btns, text="Avbryt", command=on_cancel).pack(side="right")
    ttk.Button(btns, text="Starta", command=on_ok).pack(side="right", padx=(0,8))

    toggle_time_entries()
    root.mainloop()

    if not ok_pressed["done"]:
        sys.exit(0)
    return settings

# ================== START: hämta inställningar via GUI ==================
settings = gui_collect_settings()

# Värden från GUI
input_files = settings["input_files"]
base_dir    = settings["base_dir"]
base_name   = settings["base_name"]

# --- Skapa mappen Results och vägar till Excel-utdata ---
results_dir = os.path.join(base_dir, "Results")
os.makedirs(results_dir, exist_ok=True)

out_path_nvi = os.path.join(results_dir, f"{base_name}_NVI.xlsx")
out_path_art = os.path.join(results_dir, f"{base_name}_ART.xlsx")

# --- Eventuell färg-override från GUI (endast om angiven) ---
if settings["colors"]["NVI"]["Socialt"]:       HEX_NVI_SOC   = settings["colors"]["NVI"]["Socialt"]
if settings["colors"]["NVI"]["Födosökande"]:   HEX_NVI_FODO  = settings["colors"]["NVI"]["Födosökande"]
if settings["colors"]["NVI"]["Förbiflygande"]: HEX_NVI_FORBI = settings["colors"]["NVI"]["Förbiflygande"]
if settings["colors"]["ART"]["Socialt"]:       HEX_ART_SOC   = settings["colors"]["ART"]["Socialt"]
if settings["colors"]["ART"]["Födosökande"]:   HEX_ART_FODO  = settings["colors"]["ART"]["Födosökande"]
if settings["colors"]["ART"]["Förbiflygande"]: HEX_ART_FORBI = settings["colors"]["ART"]["Förbiflygande"]

# --- (ÅTER)BYGG fyllningsscheman för tabellen med ev. nya HEX ---
SCHEME_NVI_TABLE = {
    "Socialt":       fill_from_hex(HEX_NVI_SOC),
    "Födosökande":   fill_from_hex(HEX_NVI_FODO),
    "Fodosökande":   fill_from_hex(HEX_NVI_FODO),
    "Förbiflygande": fill_from_hex(HEX_TABLE_FORBI),
}
SCHEME_ART_TABLE = {
    "Socialt":       fill_from_hex(HEX_ART_SOC),
    "Födosökande":   fill_from_hex(HEX_ART_FODO),
    "Fodosökande":   fill_from_hex(HEX_ART_FODO),
    "Förbiflygande": fill_from_hex(HEX_TABLE_FORBI),
}

# ================== STEG 1A: Läs och bygg tabellen ”Översikt” ==================
used_sheet_names = set()
sheets_to_write = []
counts_per_file = {}
total_ljud_per_file = {}
nights_per_file = {}
all_species_latin = set()

for path in input_files:
    sheet_name = safe_sheet_name(path, used_sheet_names)
    df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    sheets_to_write.append((sheet_name, df))
    total_ljud_per_file[sheet_name] = int(len(df))  # inkl. ”Noise”
    nights_per_file[sheet_name] = count_nights(df)

    if "MANUAL ID" not in df.columns:
        counts_per_file[sheet_name] = {}
        continue

    df["__list"] = df["MANUAL ID"].map(extract_species_and_type)
    long = df.explode("__list")
    long = long[long["__list"].notna()]
    if long.empty:
        counts_per_file[sheet_name] = {}
        continue

    long[["ArtLatin", "Beteendetyper"]] = pd.DataFrame(long["__list"].tolist(), index=long.index)
    mask_ok = long["ArtLatin"].astype(str).str.strip().str.lower() != "noise"
    long = long[mask_ok]

    all_species_latin.update(long["ArtLatin"].astype(str).str.strip().tolist())

    grp = long.groupby(["ArtLatin", "Beteendetyper"]).size()
    counts_per_file[sheet_name] = {(sp, typ): int(n) for (sp, typ), n in grp.items()}

# Ordning på beteendetyper i tabellen
type_order_overview = ["Socialt", "Födosökande", "Förbiflygande"]

# Artsortering (A–Ö, utom Nyctaloid/Chiroptera sist)
species_sorted_latin = sorted(all_species_latin, key=species_sort_key)

# Kolumner för olika källblad
file_cols = list(counts_per_file.keys())

# Raddata för Översikt
rows_data = []
for latin in species_sorted_latin:
    disp = display_label_multiline(latin)
    for typ in type_order_overview:
        row = {"Art": disp, "Beteendetyper": typ}
        for col in file_cols:
            val = counts_per_file.get(col, {}).get((latin, typ), 0)
            row[col] = ("" if val == 0 else int(val))
        rows_data.append(row)

# Summeringsrader
sum_row = {"Art": "", "Beteendetyper": "Fladdermusregistreringar"}
for col in file_cols:
    sum_row[col] = int(sum(counts_per_file.get(col, {}).values()))
rows_data.append(sum_row)

nights_row = {"Art": "", "Beteendetyper": "Antal nätter"}
for col in file_cols:
    n = nights_per_file.get(col)
    nights_row[col] = ("" if not n else int(n))
rows_data.append(nights_row)

per_night_row = {"Art": "", "Beteendetyper": "Antal registreringar / natt"}
for col in file_cols:
    per_night_row[col] = ""
rows_data.append(per_night_row)

tot_row = {"Art": "", "Beteendetyper": "Total antal ljud"}
for col in file_cols:
    tot_row[col] = int(total_ljud_per_file.get(col, 0))
rows_data.append(tot_row)

overview_df = pd.DataFrame(rows_data, columns=["Art", "Beteendetyper"] + file_cols)
num_species = len(species_sorted_latin)

# ================== STEG 1B: Skriv två Excel-filer och formatera ==================
def write_overview_to(path_out):
    """Skriver Översikt + kopior av alla indata-ark till en fil."""
    with pd.ExcelWriter(path_out, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Översikt", index=False)
        for sheet_name, df_orig in sheets_to_write:
            df_orig.to_excel(writer, sheet_name=sheet_name, index=False)

def format_overview(path_out, scheme_fills, num_species_rows, file_cols_list):
    """Formaterar bladet ”Översikt”: rubriker, kolumnbredder, färger, kantlinjer och formler."""
    wb = load_workbook(path_out)
    ws = wb["Översikt"]

    max_row = ws.max_row
    max_col = ws.max_column
    header_row = 1
    data_start = header_row + 1

    num_species_rows_total = num_species_rows * 3
    sum_row_idx      = data_start + num_species_rows_total
    nights_row_idx   = sum_row_idx + 1
    pernight_row_idx = nights_row_idx + 1
    total_row_idx    = pernight_row_idx + 1

    # Rubrikrad
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = FILL_HDR
        cell.font = Font(color=hex_to_argb(HEX_HEADER_FG), bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[header_row].height = 18

    # Kolumnbredder
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 24
    for idx, col_name in enumerate(file_cols_list, start=3):
        header_text = str(col_name)
        width = max(12, min(50, int(len(header_text) * 1.1)))
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Slå ihop kolumn A per artblock + radbrytning
    if num_species_rows_total > 0:
        current = data_start
        merge_end_limit = sum_row_idx - 1
        while current <= merge_end_limit:
            art_val = ws[f"A{current}"].value
            if not art_val:
                current += 1
                continue
            end = current
            while end + 1 <= merge_end_limit and ws[f"A{end+1}"].value == art_val:
                end += 1
            if end > current:
                ws.merge_cells(start_row=current, start_column=1, end_row=end, end_column=1)
            ws.cell(row=current, column=1).alignment = Alignment(vertical="center", wrap_text=True)
            current = end + 1

    # Färglägg rader efter beteendetyp
    for r in range(data_start, sum_row_idx):
        typ = ws.cell(row=r, column=2).value
        fill = scheme_fills.get(typ)
        if fill:
            ws.cell(row=r, column=2).fill = fill
            for c in range(3, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val not in (None, "", 0):
                    ws.cell(row=r, column=c).fill = fill

    # Fetstil på summeringsrader
    for r in (sum_row_idx, nights_row_idx, pernight_row_idx, total_row_idx):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).font = Font(bold=True)
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

    # Formel: ”Antal registreringar / natt”
    for col_idx in range(3, max_col + 1):
        col_letter = get_column_letter(col_idx)
        formula = f'=IFERROR({col_letter}{sum_row_idx}/{col_letter}{nights_row_idx},"")'
        cell = ws.cell(row=pernight_row_idx, column=col_idx)
        cell.value = formula
        cell.number_format = "0.0"

    # Tjock horisontell linje över varje artblock
    for i in range(num_species_rows):
        top_row = data_start + i * 3
        for c in range(1, max_col + 1):
            old = ws.cell(row=top_row, column=c).border
            ws.cell(row=top_row, column=c).border = Border(
                left=old.left, right=old.right, top=BORDER_MEDIUM, bottom=old.bottom
            )

    # Tjock linje ovanför summeringar + ram
    for c in range(1, max_col + 1):
        old = ws.cell(row=sum_row_idx, column=c).border
        ws.cell(row=sum_row_idx, column=c).border = Border(
            left=old.left, right=old.right, top=BORDER_MEDIUM, bottom=old.bottom
        )
        old = ws.cell(row=total_row_idx, column=c).border
        ws.cell(row=total_row_idx, column=c).border = Border(
            left=old.left, right=BORDER_MEDIUM, top=old.top, bottom=BORDER_MEDIUM
        )

    # Tjocka vertikala avgränsningar
    for c in range(3, max_col + 1):
        for r in range(header_row, max_row + 1):
            old = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(
                left=BORDER_MEDIUM, right=old.right, top=old.top, bottom=old.bottom
            )
    for r in range(header_row, max_row + 1):
        oldA = ws.cell(row=r, column=1).border
        ws.cell(row=r, column=1).border = Border(
            left=BORDER_MEDIUM, right=oldA.right, top=oldA.top, bottom=oldA.bottom
        )
        oldB = ws.cell(row=r, column=2).border
        ws.cell(row=r, column=2).border = Border(
            left=BORDER_MEDIUM, right=oldB.right, top=oldB.top, bottom=oldB.bottom
        )
        oldR = ws.cell(row=r, column=max_col).border
        ws.cell(row=r, column=max_col).border = Border(
            left=oldR.left, right=BORDER_MEDIUM, top=oldR.top, bottom=oldR.bottom
        )

    wb.save(path_out)
    wb.close()

# Skriv och formatera NVI-tabell
write_overview_to(out_path_nvi)
format_overview(out_path_nvi, SCHEME_NVI_TABLE, num_species, file_cols)
print(f"Klar! Sparad fil (NVI): {out_path_nvi}")

# Skriv och formatera ART-tabell
write_overview_to(out_path_art)
format_overview(out_path_art, SCHEME_ART_TABLE, num_species, file_cols)
print(f"Klar! Sparad fil (ART): {out_path_art}")

# Öppna båda för snabb kontroll
open_file(out_path_nvi)
open_file(out_path_art)

# ================== STEG 2: (valfritt) Skapa diagram – global Y-skala ==================
def _hm_from_any_for_plot(val):
    """Alias för tidsparsing i plottlogik (samma funktion som _hm_from_any)."""
    return _hm_from_any(val)

def _compute_file_ymax(input_file, custom_time_range):
    """Beräknar max staplad topp (per 15-min slot) för en fil – används för global Y-skala."""
    try:
        df = pd.read_excel(input_file)
        df["species_type_list"] = df["MANUAL ID"].map(extract_species_and_type)

        def time_to_interval(val):
            hm = _hm_from_any_for_plot(val)
            if hm is None:
                return ""
            h, m = hm
            minutes = int((m // 15) * 15)
            return f"{h:02d}:{minutes:02d}"
        df["interval"] = df["TIME"].map(time_to_interval)

        df_long = df.explode("species_type_list")
        df_long = df_long[df_long["species_type_list"].notna()]
        if df_long.empty:
            return 0
        df_long[["species", "obs_type"]] = pd.DataFrame(df_long["species_type_list"].tolist(), index=df_long.index)
        df_long = df_long[df_long["species"].astype(str).str.strip().str.lower() != "noise"]
        df_long = df_long[df_long["species"].astype(str).str.strip() != ""]

        if custom_time_range:
            min_dt = str_to_dt(custom_time_range[0] + ":00")
            max_dt = str_to_dt(custom_time_range[1] + ":00")
            if min_dt is None or max_dt is None:
                return 0
            min_dt = round_down_15(min_dt)
            max_dt = round_up_15(max_dt)
        else:
            dt_series = df["TIME"].apply(str_to_dt).dropna()
            if len(dt_series) == 0:
                ints = [s for s in df["interval"].astype(str).tolist() if s and s.lower() != "nan"]
                dt_from_int = [interval_to_sortkey(s) for s in ints]
                dt_from_int = [d for d in dt_from_int if d is not None]
                if not dt_from_int:
                    return 0
                min_dt = round_down_15(min(dt_from_int))
                max_dt = round_up_15(max(dt_from_int))
            else:
                min_dt = round_down_15(min(dt_series))
                max_dt = round_up_15(max(dt_series))

        all_intervals = []
        t = min_dt
        while t <= max_dt:
            all_intervals.append(t.strftime("%H:%M"))
            t += timedelta(minutes=15)
        all_intervals = list(dict.fromkeys(all_intervals))

        agg = df_long.groupby(["interval", "species", "obs_type"]).size().reset_index(name="antal")
        agg["interval"] = pd.Categorical(agg["interval"], categories=all_intervals, ordered=True)
        type_order = ["Socialt", "Födosökande", "Förbiflygande"]

        y_max_file = 0
        for sp in df_long["species"].unique():
            plot_data = (
                agg[agg["species"] == sp]
                .pivot(index="interval", columns="obs_type", values="antal")
                .fillna(0)
                .reindex(all_intervals, fill_value=0)
                .reindex(columns=type_order, fill_value=0)
            )
            if not plot_data.empty:
                y_max_file = max(y_max_file, int(plot_data.sum(axis=1).max()))
        return y_max_file
    except Exception:
        return 0

def generate_bat_diagrams(input_files_list, diagrams_root, custom_time_range):
    """Genererar linje- och stapeldiagram för varje fil – alla med gemensam global Y-skala."""
    # Globalt Y-tak baserat på alla valda filer
    global_y_max = 0
    for p in input_files_list:
        global_y_max = max(global_y_max, _compute_file_ymax(p, custom_time_range))
    y_lim_global = max(1, math.ceil(global_y_max * 1.05))
    print(f"Global gemensam Y-max (linje + stapel) för alla filer: {y_lim_global}")

    for input_file in input_files_list:
        print(f"\nBearbetar: {os.path.basename(input_file)}")

        stem = os.path.splitext(os.path.basename(input_file))[0]
        output_dir_lines  = os.path.join(diagrams_root, f"{stem}_linjediagram")
        output_dir_stacks = os.path.join(diagrams_root, f"{stem}_stapeldiagram")
        os.makedirs(output_dir_lines, exist_ok=True)

        # Läs och förbered
        df = pd.read_excel(input_file)
        df["species_type_list"] = df["MANUAL ID"].map(extract_species_and_type)

        def time_to_interval(val):
            hm = _hm_from_any_for_plot(val)
            if hm is None:
                return ""
            h, m = hm
            minutes = int((m // 15) * 15)
            return f"{h:02d}:{minutes:02d}"
        df["interval"] = df["TIME"].map(time_to_interval)

        df_long = df.explode("species_type_list")
        df_long = df_long[df_long["species_type_list"].notna()]
        if df_long.empty:
            print("Inga data efter tolkning. Hoppar över.")
            continue
        df_long[["species", "obs_type"]] = pd.DataFrame(df_long["species_type_list"].tolist(), index=df_long.index)
        df_long = df_long[df_long["species"].astype(str).str.strip().str.lower() != "noise"]
        df_long = df_long[df_long["species"].astype(str).str.strip() != ""]

        # Tidsspann (auto eller manuellt)
        if custom_time_range:
            min_dt = str_to_dt(custom_time_range[0] + ":00")
            max_dt = str_to_dt(custom_time_range[1] + ":00")
            if min_dt is None or max_dt is None:
                print("Fel i manuellt intervall. Hoppar över filen.")
                continue
            min_dt = round_down_15(min_dt)
            max_dt = round_up_15(max_dt)
        else:
            dt_series = df["TIME"].apply(str_to_dt).dropna()
            if len(dt_series) == 0:
                ints = [s for s in df["interval"].astype(str).tolist() if s and s.lower() != "nan"]
                dt_from_int = [interval_to_sortkey(s) for s in ints]
                dt_from_int = [d for d in dt_from_int if d is not None]
                if not dt_from_int:
                    print("Inga giltiga tider. Hoppar över filen.")
                    continue
                min_dt = round_down_15(min(dt_from_int))
                max_dt = round_up_15(max(dt_from_int))

            else:
                min_dt = round_down_15(min(dt_series))
                max_dt = round_up_15(max(dt_series))

        # Lista alla 15-minutersintervall
        all_intervals = []
        t = min_dt
        while t <= max_dt:
            all_intervals.append(t.strftime("%H:%M"))
            t += timedelta(minutes=15)
        all_intervals = list(dict.fromkeys(all_intervals))

        # Aggregera data för diagram
        agg = df_long.groupby(["interval", "species", "obs_type"]).size().reset_index(name="antal")
        agg["interval"] = pd.Categorical(agg["interval"], categories=all_intervals, ordered=True)
        species_list = sorted(df_long["species"].unique())
        type_order = ["Socialt", "Födosökande", "Förbiflygande"]

        # Färglistor (HEX) för stapeldiagram
        colors_art = [HEX_ART_SOC, HEX_ART_FODO, HEX_ART_FORBI]
        colors_nvi = [HEX_NVI_SOC, HEX_NVI_FODO, HEX_NVI_FORBI]

        # LINJEDIAGRAM – samlingsdiagram
        agg_line = df_long.groupby(["interval", "species"]).size().reset_index(name="antal")
        agg_line["interval"] = pd.Categorical(agg_line["interval"], categories=all_intervals, ordered=True)
        pivot_line = agg_line.pivot(index="interval", columns="species", values="antal").fillna(0)
        pivot_line = pivot_line.reindex(all_intervals, fill_value=0)
        total_obs = df_long.shape[0]

        plt.figure(figsize=(12, 6))
        ax_all = plt.gca()
        pivot_line.plot(ax=ax_all, marker='o')
        ax_all.set_xlabel("Tid (15-minutersintervall)")
        ax_all.set_ylabel("Antal ljudfiler")
        ax_all.set_title(f"Fladdermusobservationer – alla arter, antal observerade beteenden: {total_obs}")
        ax_all.legend(title="Art", bbox_to_anchor=(1.05, 1), loc='upper left')
        ax_all.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax_all.set_xticks(range(len(pivot_line.index)))
        ax_all.set_xticklabels(pivot_line.index, rotation=270)
        ax_all.set_ylim(0, y_lim_global)
        plt.tight_layout()
        plt.grid(True, axis='y')
        plt.savefig(os.path.join(output_dir_lines, "alla_arter.png"))
        plt.close()

        # LINJEDIAGRAM – per art
        for species in species_list:
            total_sp = int(pivot_line[species].sum())
            plt.figure(figsize=(10, 4))
            ax = plt.gca()
            ax.plot(pivot_line.index, pivot_line[species], marker='o')
            ax.set_xlabel("Tid (15-minutersintervall)")
            ax.set_ylabel("Antal ljudfiler")
            ax.set_title(format_title(species, total_sp))
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            ax.set_xticks(range(len(pivot_line.index)))
            ax.set_xticklabels(pivot_line.index, rotation=270)
            ax.set_ylim(0, y_lim_global)
            plt.grid(True, axis='y')
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir_lines, f"{safe_filename(species)}.png"))
            plt.close()

        # STAPELDIAGRAM – ART
        output_dir_stacks_art = output_dir_stacks + "_ART"
        os.makedirs(output_dir_stacks_art, exist_ok=True)
        for species in species_list:
            plot_data = (
                agg[agg["species"] == species]
                .pivot(index="interval", columns="obs_type", values="antal")
                .fillna(0)
                .reindex(all_intervals, fill_value=0)
                .reindex(columns=type_order, fill_value=0)
            )
            ax = plot_data.plot(kind="bar", stacked=True, color=colors_art, figsize=(14, 6))
            plt.xlabel("Tid (15-minutersintervall)")
            plt.ylabel("Antal ljudfiler")
            plt.title(format_title(species, int(plot_data.values.sum())))
            plt.legend(title="Beteendetyper")
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            ax.set_ylim(0, y_lim_global)
            plt.xticks(rotation=270)
            plt.tight_layout()
            plt.grid(True, axis='y')
            plt.savefig(os.path.join(output_dir_stacks_art, f"{safe_filename(species)}.png"))
            plt.close()
        print(f"Stapeldiagram ART sparade i mappen: {output_dir_stacks_art}")

        # STAPELDIAGRAM – NVI
        output_dir_stacks_nvi = output_dir_stacks + "_NVI"
        os.makedirs(output_dir_stacks_nvi, exist_ok=True)
        for species in species_list:
            plot_data = (
                agg[agg["species"] == species]
                .pivot(index="interval", columns="obs_type", values="antal")
                .fillna(0)
                .reindex(all_intervals, fill_value=0)
                .reindex(columns=type_order, fill_value=0)
            )
            ax = plot_data.plot(kind="bar", stacked=True, color=colors_nvi, figsize=(14, 6))
            plt.xlabel("Tid (15-minutersintervall)")
            plt.ylabel("Antal ljudfiler")
            plt.title(format_title(species, int(plot_data.values.sum())))
            plt.legend(title="Beteendetyper")
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            ax.set_ylim(0, y_lim_global)
            plt.xticks(rotation=270)
            plt.tight_layout()
            plt.grid(True, axis='y')
            plt.savefig(os.path.join(output_dir_stacks_nvi, f"{safe_filename(species)}.png"))
            plt.close()
        print(f"Stapeldiagram NVI sparade i mappen: {output_dir_stacks_nvi}")
        print(f"Linjediagram sparade i mappen: {output_dir_lines}")
        print(f"Tidsintervall: {all_intervals[0]} – {all_intervals[-1]}")

# ================== Kör diagram om valt i GUI ==================
if settings["do_plots"]:
    diagrams_base = settings["diagram_base"] or results_dir
    diagrams_root = os.path.join(diagrams_base, "diagramer")
    os.makedirs(diagrams_root, exist_ok=True)
    print(f"Resultat kommer att sparas i: {diagrams_root}")

    custom_time_range = settings["custom_time_range"]  # None → auto
    generate_bat_diagrams(input_files, diagrams_root, custom_time_range)
