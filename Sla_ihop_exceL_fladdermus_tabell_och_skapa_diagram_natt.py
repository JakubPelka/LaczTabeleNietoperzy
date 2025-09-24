# Rekommenderat filnamn:
# Sla_ihop_exceL_fladdermus_tabell_och_skapa_diagram_dobowy.py
#
# Nytt:
# - GUI z dwoma niezależnymi opcjami: „Generera samlade diagram” i „Generera natt-för-natt diagram”.
# - Nocne wykresy tworzone DODATKOWO do standardowych.
# - Etykieta nocy w tytułach: DD/NN.MM (np. 22/23.08).
# Reszta logiki (Excel *_NVI i *_ART, globalna oś Y, kolory, X-oś auto/manual) bez zmian.

import os
import re
import sys
import math
import subprocess
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta, date
from matplotlib.ticker import MaxNLocator

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ================== Grundinställningar / etiketter ==================
DEFAULT_OUT_BASENAME = "sammanstallning_fladdermus"

# ================== Färger – enhetligt HEX-format ==================
HEX_HEADER_BG = "#595959"
HEX_HEADER_FG = "#FFFFFF"
HEX_TABLE_FORBI = "#E7E6E6"   # samma i NVI/ART tabellerna

# Stapeldiagram-paletter (standard)
HEX_NVI_SOC   = "#FF0000"  # Socialt
HEX_NVI_FODO  = "#FFC000"  # Födosökande
HEX_NVI_FORBI = "#A9A9A9"  # Förbiflygande

HEX_ART_SOC   = "#EB09D8"  # Socialt
HEX_ART_FODO  = "#D98FD3"  # Födosökande
HEX_ART_FORBI = "#ABAAA9"  # Förbiflygande

BORDER_MEDIUM = Side(style="medium", color="FF000000")

def hex_to_argb(hex_rgb: str) -> str:
    h = hex_rgb.strip().lstrip("#")
    if len(h) != 6:
        raise ValueError(f"Ogiltig HEX: {hex_rgb}")
    return "FF" + h.upper()

def fill_from_hex(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_to_argb(hex_rgb))

FILL_HDR = fill_from_hex(HEX_HEADER_BG)

# =============== Ordbok Latin → Svenska ===============
LATIN_TO_SV = {
    "Barbastella barbastellus": "Barbastell",
    "Eptesicus nilssonii": "Nordfladdermus",
    "Eptesicus serotinus": "Sydfladdermus",
    "Myotis alcathoe": "Nymffladdermus",
    "Myotis bechsteinii": "Bechsteins fladdermus",
    "Myotis brandtii": "Tajgafladdermus",
    "Myotis dasycneme": "Dammfladdermus",
    "Myotis daubentonii": "Vattenfladdermus",
    "Myotis myotis": "Större musöra",
    "Myotis mystacinus": "Mustaschfladdermus",
    "Myotis nattereri": "Fransfladdermus",
    "Myotis mystacinus/brandtii": "Mustasch/Tajgafladdermus",
    "Nyctalus leisleri": "Mindre brunfladdermus",
    "Nyctalus noctula": "Större brunfladdermus",
    "Pipistrellus kuhlii": "Parkpipistrell",
    "Pipistrellus nathusii": "Trollpipistrell",
    "Pipistrellus pipistrellus": "Sydpipistrell",
    "Pipistrellus pygmaeus": "Dvärgpipistrell",
    "Plecotus auritus": "Brunlångöra",
    "Plecotus austriacus": "Grålångöra",
    "Vespertilio murinus": "Gråskimlig fladdermus",
    "Nyctaloid": None,
    "Chiroptera": None,
}
SPECIAL_TAIL = {"nyctaloid", "chiroptera"}  # sorteras sist i artlistor

# ================== Hjälpfunktioner (gemensamma) ==================
def safe_sheet_name(path: str, used: set) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    base = re.sub(r'[:\\\/\?\*\[\]]', '_', base).strip()[:31] or "Ark"
    cand = base; i = 2
    while cand in used or not cand:
        suf = f"_{i}"
        cand = (base[: (31 - len(suf))] + suf) if len(base) + len(suf) > 31 else (base + suf)
        i += 1
    used.add(cand); return cand

def display_label_multiline(latin: str) -> str:
    latin = str(latin).strip()
    sv = LATIN_TO_SV.get(latin, None)
    if sv:
        return f"{sv.capitalize()},\n{latin}"
    return latin

def format_title(species_latin: str, total_count: int, night_label: str | None = None) -> str:
    sv = LATIN_TO_SV.get(species_latin)
    base = f"{sv} ({species_latin})" if sv else species_latin
    if night_label:
        return f"{base} – natt {night_label}, antal observerade beteenden: {int(total_count)}"
    return f"{base}, antal observerade beteenden: {int(total_count)}"

def extract_species_and_type(manual_id_value):
    if pd.isna(manual_id_value): return []
    out = []
    for raw in str(manual_id_value).split(","):
        entry = raw.strip()
        if not entry: continue
        upper = entry.upper()
        if "FOD" in upper: typ = "Födosökande"
        elif "SOC" in upper: typ = "Socialt"
        else: typ = "Förbiflygande"
        species = re.sub(r'\b(FOD|SOC)\b', "", entry, flags=re.IGNORECASE).strip(" ,;")
        out.append((species, typ))
    return out

def open_file(path):
    try:
        if sys.platform.startswith("win"): os.startfile(path)
        elif sys.platform == "darwin": subprocess.run(["open", path])
        else: subprocess.run(["xdg-open", path])
    except Exception as e:
        print(f"Kan inte öppna filen automatiskt: {e}")

def species_sort_key(latin: str):
    return (str(latin).strip().lower() in SPECIAL_TAIL, str(latin).casefold())

def safe_filename(s):
    return re.sub(r'[\\/:\*\?"<>\|]', '_', str(s))

# --- Tidshjälp ---
def _hm_from_any(val):
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)) or (isinstance(val, str) and val.strip() == ""):
            return None
        if hasattr(val, "hour") and hasattr(val, "minute"):
            return int(val.hour), int(val.minute)
        if isinstance(val, (int, float)) and not pd.isna(val):
            frac = float(val) % 1.0
            secs = int(round(frac * 24 * 60 * 60))
            h = (secs // 3600) % 24
            m = (secs % 3600) // 60
            return int(h), int(m)
        s = str(val).strip()
        t = pd.to_datetime(s, format="%H:%M:%S", errors="coerce")
        if pd.isna(t):
            t = pd.to_datetime(s, format="%H:%M", errors="coerce")
        if pd.isna(t):
            return None
        return int(t.hour), int(t.minute)
    except Exception:
        return None

def str_to_dt(time_val):
    hm = _hm_from_any(time_val)
    if hm is None:
        return None
    h, m = hm
    fake_date = "2000-01-02" if h < 12 else "2000-01-01"
    return datetime.strptime(f"{fake_date} {h:02d}:{m:02d}", "%Y-%m-%d %H:%M")

def round_down_15(dt):
    return dt.replace(minute=(dt.minute // 15) * 15, second=0, microsecond=0)

def round_up_15(dt):
    if dt.minute % 15 != 0 or dt.second > 0 or dt.microsecond > 0:
        dt = dt + timedelta(minutes=15 - (dt.minute % 15), seconds=-dt.second, microseconds=-dt.microsecond)
    return dt.replace(second=0, microsecond=0)

def interval_to_sortkey(interval):
    try:
        t = pd.to_datetime(str(interval), format="%H:%M", errors="coerce")
        if pd.isna(t): return None
        h, m = int(t.hour), int(t.minute)
        fake_date = "2000-01-02" if h < 12 else "2000-01-01"
        return datetime.strptime(f"{fake_date} {h:02d}:{m:02d}", "%Y-%m-%d %H:%M")
    except Exception:
        return None

def detect_column(df, candidates):
    lowmap = {str(c).strip().lower(): c for c in df.columns}
    for k in candidates:
        if k in lowmap:
            return lowmap[k]
    return None

def count_nights(df):
    date_col = detect_column(df, ["date", "datum"])
    if not date_col: return None
    time_col = detect_column(df, ["time", "tid"])

    dates = pd.to_datetime(df[date_col], errors="coerce")
    if time_col:
        hm = df[time_col].apply(_hm_from_any)
        hours = hm.apply(lambda x: x[0] if isinstance(x, tuple) else None)
        shift = hours.apply(lambda h: (h is not None) and (h < 12))
        night_key = (dates.dt.normalize() - pd.to_timedelta(shift.fillna(False).astype(int), unit="D")).dt.date
    else:
        night_key = dates.dt.normalize().dt.date

    nights = pd.Series(night_key).dropna().nunique()
    return int(nights) if nights else None

def row_night_key(d_val, t_val):
    d = pd.to_datetime(d_val, errors="coerce")
    if pd.isna(d): return None
    hm = _hm_from_any(t_val)
    if hm is None:
        return d.date()
    h, _ = hm
    return (d - pd.Timedelta(days=1)).date() if h < 12 else d.date()

# ================== GUI ==================
def _validate_hex(s):
    s = (s or "").strip()
    if not s: return None
    if s.startswith("#"): s = s[1:]
    if len(s) != 6 or any(c not in "0123456789abcdefABCDEF" for c in s):
        return None
    return "#" + s.upper()

def gui_collect_settings(default_basename=DEFAULT_OUT_BASENAME):
    root = tk.Tk()
    root.title("Fladdermus – sammanställning & diagram (natt/löpande)")
    root.geometry("920x800"); root.minsize(860, 680)
    main = ttk.Frame(root, padding=12); main.pack(fill="both", expand=True)

    # Indata
    lf_in = ttk.LabelFrame(main, text="Indatafiler"); lf_in.pack(fill="both", expand=False, pady=(0,10))
    files_list = tk.Listbox(lf_in, height=7, selectmode="extended")
    files_list.grid(row=0, column=0, rowspan=3, sticky="nsew", padx=(8,8), pady=8)
    lf_in.columnconfigure(0, weight=1); lf_in.rowconfigure(0, weight=1)
    ttk.Button(lf_in, text="Lägg till filer…",
               command=lambda: [files_list.insert(tk.END, p) for p in filedialog.askopenfilenames(
                   title="Välj Excel-filer", filetypes=[("Excel", "*.xlsx"), ("Alla", "*.*")]
               ) or []]).grid(row=0, column=1, sticky="ew", padx=(0,8), pady=(8,4))
    ttk.Button(lf_in, text="Rensa listan", command=lambda: files_list.delete(0, tk.END))\
        .grid(row=1, column=1, sticky="ew", padx=(0,8))

    # Utdata + basnamn
    lf_out = ttk.LabelFrame(main, text="Utdatakatalog och basnamn"); lf_out.pack(fill="x", expand=False, pady=(0,10))
    var_outdir = tk.StringVar(value=""); var_basename = tk.StringVar(value=default_basename)
    ttk.Label(lf_out, text="Basnamn (utan ändelse):").grid(row=0, column=0, sticky="w", padx=8, pady=(8,4))
    ttk.Entry(lf_out, textvariable=var_basename).grid(row=0, column=1, sticky="ew", padx=8, pady=(8,4))
    ttk.Label(lf_out, text="Mapp för 'Results/':").grid(row=1, column=0, sticky="w", padx=8)
    ttk.Entry(lf_out, textvariable=var_outdir).grid(row=1, column=1, sticky="ew", padx=8)
    ttk.Button(lf_out, text="Välj mapp…",
               command=lambda: var_outdir.set(filedialog.askdirectory(title="Välj mapp") or var_outdir.get()))\
        .grid(row=1, column=2, sticky="ew", padx=(0,8))
    lf_out.columnconfigure(1, weight=1)

    # Diagram (dwie niezależne opcje)
    lf_plot = ttk.LabelFrame(main, text="Diagram (valfritt)"); lf_plot.pack(fill="x", expand=False, pady=(0,10))
    var_do_plots_summary = tk.BooleanVar(value=True)
    var_do_plots_pernight = tk.BooleanVar(value=True)
    ttk.Checkbutton(lf_plot, text="Generera samlade diagram", variable=var_do_plots_summary)\
        .grid(row=0, column=0, sticky="w", padx=8, pady=(8,4))
    ttk.Checkbutton(lf_plot, text="Generera natt-för-natt diagram", variable=var_do_plots_pernight)\
        .grid(row=0, column=1, sticky="w", padx=8, pady=(8,4))

    var_time_mode = tk.StringVar(value="auto")
    var_tstart = tk.StringVar(value="22:00"); var_tend = tk.StringVar(value="02:00")
    ttk.Radiobutton(lf_plot, text="X-axel: automatisk (från data)", value="auto", variable=var_time_mode)\
        .grid(row=1, column=0, sticky="w", padx=8)
    ttk.Radiobutton(lf_plot, text="X-axel: eget intervall", value="manual", variable=var_time_mode)\
        .grid(row=1, column=1, sticky="w")
    ttk.Label(lf_plot, text="Starttid (HH:MM):").grid(row=2, column=0, sticky="e", padx=8, pady=(4,8))
    ent_start = ttk.Entry(lf_plot, textvariable=var_tstart, width=10); ent_start.grid(row=2, column=1, sticky="w", pady=(4,8))
    ttk.Label(lf_plot, text="Sluttid (HH:MM):").grid(row=2, column=2, sticky="e", padx=8, pady=(4,8))
    ent_end = ttk.Entry(lf_plot, textvariable=var_tend, width=10); ent_end.grid(row=2, column=3, sticky="w", pady=(4,8))

    ttk.Label(lf_plot, text="Bas-mapp för diagram (valfritt, default Results/diagramer):").grid(row=3, column=0, sticky="e", padx=8, pady=(0,8))
    var_diagdir = tk.StringVar(value="")
    ttk.Entry(lf_plot, textvariable=var_diagdir).grid(row=3, column=1, sticky="ew", padx=8, pady=(0,8), columnspan=2)
    ttk.Button(lf_plot, text="Välj…", command=lambda: var_diagdir.set(filedialog.askdirectory(title="Välj basmapp") or var_diagdir.get()))\
        .grid(row=3, column=3, sticky="ew", pady=(0,8))
    lf_plot.columnconfigure(1, weight=1)

    # Färger + paleta
    lf_colors = ttk.LabelFrame(main, text="Färger (valfritt – lämna tomt för standard)"); lf_colors.pack(fill="x", expand=False)
    def bind_preview(var, preview_widget):
        def _cb(*_):
            hx = _validate_hex(var.get())
            preview_widget.config(background=(hx or "#FFFFFF"))
        var.trace_add("write", _cb); _cb()
    def color_row(row, label, vars_tuple):
        ttk.Label(lf_colors, text=label).grid(row=row, column=0, sticky="w", padx=8, pady=(8,2))
        def one(col_ix, var):
            e = ttk.Entry(lf_colors, textvariable=var, width=10)
            e.grid(row=row, column=1 + col_ix*3, padx=(4,2), pady=(8,2), sticky="w")
            prev = tk.Label(lf_colors, text="  ", width=3, relief="groove")
            prev.grid(row=row, column=2 + col_ix*3, padx=(0,2), pady=(8,2), sticky="w")
            bind_preview(var, prev)
            def choose():
                init = _validate_hex(var.get()) or "#FFFFFF"
                _, hx = colorchooser.askcolor(color=init, title="Välj färg")
                if hx: var.set(hx.upper())
            ttk.Button(lf_colors, text="Välj…", command=choose)\
                .grid(row=row, column=3 + col_ix*3, padx=(0,6), pady=(8,2), sticky="w")
        for i, v in enumerate(vars_tuple): one(i, v)
    var_nvi_soc=tk.StringVar(value="#FF0000"); var_nvi_fodo=tk.StringVar(value="#FFC000"); var_nvi_forbi=tk.StringVar(value="#A9A9A9")
    var_art_soc=tk.StringVar(value="#EB09D8"); var_art_fodo=tk.StringVar(value="#D98FD3"); var_art_forbi=tk.StringVar(value="#ABAAA9")
    color_row(0, "NVI – Socialt / Födosökande / Förbiflygande:", (var_nvi_soc, var_nvi_fodo, var_nvi_forbi))
    color_row(1, "ART – Socialt / Födosökande / Förbiflygande:", (var_art_soc, var_art_fodo, var_art_forbi))
    for c in range(10): lf_colors.columnconfigure(c, weight=0)
    lf_colors.columnconfigure(9, weight=1)

    # Knappfält
    btns = ttk.Frame(main); btns.pack(fill="x", pady=(12,0))
    settings = {}; done = {"ok": False}
    def on_ok():
        files = list(files_list.get(0, tk.END))
        if not files:
            messagebox.showwarning("GUI", "Välj minst en indatafil."); return
        outdir = var_outdir.get().strip()
        if not outdir:
            messagebox.showwarning("GUI", "Välj mapp där 'Results/' ska skapas."); return
        basename = var_basename.get().strip() or DEFAULT_OUT_BASENAME

        # tryb czasu
        tmode = var_time_mode.get()
        tstart, tend = var_tstart.get().strip(), var_tend.get().strip()
        custom_range = None
        if tmode == "manual":
            try:
                pd.to_datetime(tstart, format="%H:%M"); pd.to_datetime(tend, format="%H:%M")
            except Exception:
                messagebox.showerror("GUI", "Felaktigt tidsformat. Använd HH:MM."); return
            custom_range = (tstart, tend)

        settings.update({
            "input_files": files,
            "base_dir": outdir,
            "base_name": basename,
            "custom_time_range": custom_range,   # None = auto
            "diagram_base": (var_diagdir.get().strip() or None),
            "do_plots_summary": bool(var_do_plots_summary.get()),
            "do_plots_pernight": bool(var_do_plots_pernight.get()),
            "colors": {
                "NVI": {"Socialt": _validate_hex(var_nvi_soc.get()), "Födosökande": _validate_hex(var_nvi_fodo.get()), "Förbiflygande": _validate_hex(var_nvi_forbi.get())},
                "ART": {"Socialt": _validate_hex(var_art_soc.get()), "Födosökande": _validate_hex(var_art_fodo.get()), "Förbiflygande": _validate_hex(var_art_forbi.get())},
            }
        })
        done["ok"] = True; root.destroy()
    def on_cancel():
        root.destroy(); sys.exit(0)
    ttk.Button(btns, text="Avbryt", command=on_cancel).pack(side="right")
    ttk.Button(btns, text="Starta", command=on_ok).pack(side="right", padx=(0,8))
    root.mainloop()
    if not done["ok"]: sys.exit(0)
    return settings

# ================== Start: hämta inställningar ==================
settings = gui_collect_settings()

# Indata/utdata
input_files = settings["input_files"]
base_dir    = settings["base_dir"]
base_name   = settings["base_name"]

results_dir = os.path.join(base_dir, "Results")
os.makedirs(results_dir, exist_ok=True)

out_path_nvi = os.path.join(results_dir, f"{base_name}_NVI.xlsx")
out_path_art = os.path.join(results_dir, f"{base_name}_ART.xlsx")

# Eventuell färg-override
if settings["colors"]["NVI"]["Socialt"]:       HEX_NVI_SOC   = settings["colors"]["NVI"]["Socialt"]
if settings["colors"]["NVI"]["Födosökande"]:   HEX_NVI_FODO  = settings["colors"]["NVI"]["Födosökande"]
if settings["colors"]["NVI"]["Förbiflygande"]: HEX_NVI_FORBI = settings["colors"]["NVI"]["Förbiflygande"]
if settings["colors"]["ART"]["Socialt"]:       HEX_ART_SOC   = settings["colors"]["ART"]["Socialt"]
if settings["colors"]["ART"]["Födosökande"]:   HEX_ART_FODO  = settings["colors"]["ART"]["Födosökande"]
if settings["colors"]["ART"]["Förbiflygande"]: HEX_ART_FORBI = settings["colors"]["ART"]["Förbiflygande"]

def scheme_table_nvi():
    return {
        "Socialt":       fill_from_hex(HEX_NVI_SOC),
        "Födosökande":   fill_from_hex(HEX_NVI_FODO),
        "Fodosökande":   fill_from_hex(HEX_NVI_FODO),
        "Förbiflygande": fill_from_hex(HEX_TABLE_FORBI),
    }
def scheme_table_art():
    return {
        "Socialt":       fill_from_hex(HEX_ART_SOC),
        "Födosökande":   fill_from_hex(HEX_ART_FODO),
        "Fodosökande":   fill_from_hex(HEX_ART_FODO),
        "Förbiflygande": fill_from_hex(HEX_TABLE_FORBI),
    }

# ================== STEG 1: Bygg Excel-översikt (identiskt) ==================
used_sheet_names = set()
sheets_to_write = []
counts_per_file = {}
total_ljud_per_file = {}
nights_per_file = {}
all_species_latin = set()

for path in input_files:
    sheet_name = safe_sheet_name(path, used_sheet_names)
    df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    sheets_to_write.append((sheet_name, df))
    total_ljud_per_file[sheet_name] = int(len(df))  # inkl. Noise
    nights_per_file[sheet_name] = count_nights(df)

    if "MANUAL ID" not in df.columns:
        counts_per_file[sheet_name] = {}
        continue

    df["__list"] = df["MANUAL ID"].map(extract_species_and_type)
    long = df.explode("__list"); long = long[long["__list"].notna()]
    if long.empty:
        counts_per_file[sheet_name] = {}
        continue

    long[["ArtLatin", "Beteendetyper"]] = pd.DataFrame(long["__list"].tolist(), index=long.index)
    long = long[long["ArtLatin"].astype(str).str.strip().str.lower() != "noise"]
    all_species_latin.update(long["ArtLatin"].astype(str).str.strip().tolist())
    grp = long.groupby(["ArtLatin", "Beteendetyper"]).size()
    counts_per_file[sheet_name] = {(sp, typ): int(n) for (sp, typ), n in grp.items()}

type_order_overview = ["Socialt", "Födosökande", "Förbiflygande"]
species_sorted_latin = sorted(all_species_latin, key=species_sort_key)
file_cols = list(counts_per_file.keys())

rows_data = []
for latin in species_sorted_latin:
    disp = display_label_multiline(latin)
    for typ in type_order_overview:
        row = {"Art": disp, "Beteendetyper": typ}
        for col in file_cols:
            val = counts_per_file.get(col, {}).get((latin, typ), 0)
            row[col] = ("" if val == 0 else int(val))
        rows_data.append(row)

sum_row = {"Art": "", "Beteendetyper": "Fladdermusregistreringar"}
for col in file_cols: sum_row[col] = int(sum(counts_per_file.get(col, {}).values()))
rows_data.append(sum_row)

nights_row = {"Art": "", "Beteendetyper": "Antal nätter"}
for col in file_cols:
    n = nights_per_file.get(col)
    nights_row[col] = ("" if not n else int(n))
rows_data.append(nights_row)

per_night_row = {"Art": "", "Beteendetyper": "Antal registreringar / natt"}
for col in file_cols: per_night_row[col] = ""
rows_data.append(per_night_row)

tot_row = {"Art": "", "Beteendetyper": "Total antal ljud"}
for col in file_cols: tot_row[col] = int(total_ljud_per_file.get(col, 0))
rows_data.append(tot_row)

overview_df = pd.DataFrame(rows_data, columns=["Art", "Beteendetyper"] + file_cols)
num_species = len(species_sorted_latin)

def write_overview_to(path_out):
    with pd.ExcelWriter(path_out, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Översikt", index=False)
        for sheet_name, df_orig in sheets_to_write:
            df_orig.to_excel(writer, sheet_name=sheet_name, index=False)

def format_overview(path_out, scheme_fills, num_species_rows, file_cols_list):
    wb = load_workbook(path_out); ws = wb["Översikt"]
    max_row = ws.max_row; max_col = ws.max_column
    header_row = 1; data_start = header_row + 1
    num_species_rows_total = num_species_rows * 3
    sum_row_idx      = data_start + num_species_rows_total
    nights_row_idx   = sum_row_idx + 1
    pernight_row_idx = nights_row_idx + 1
    total_row_idx    = pernight_row_idx + 1

    # rubriker
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = FILL_HDR; cell.font = Font(color=hex_to_argb(HEX_HEADER_FG), bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[header_row].height = 18

    # kolumnbredder
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 24
    for idx, col_name in enumerate(file_cols_list, start=3):
        header_text = str(col_name)
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(50, int(len(header_text) * 1.1)))

    # slå ihop artetiketter (block om 3 rader)
    if num_species_rows_total > 0:
        current = data_start; merge_end_limit = sum_row_idx - 1
        while current <= merge_end_limit:
            art_val = ws[f"A{current}"].value
            if not art_val: current += 1; continue
            end = current
            while end + 1 <= merge_end_limit and ws[f"A{end+1}"].value == art_val:
                end += 1
            if end > current:
                ws.merge_cells(start_row=current, start_column=1, end_row=end, end_column=1)
            ws.cell(row=current, column=1).alignment = Alignment(vertical="center", wrap_text=True)
            current = end + 1

    # färg rader per typ
    for r in range(data_start, sum_row_idx):
        typ = ws.cell(row=r, column=2).value
        fill = scheme_fills.get(typ)
        if fill:
            ws.cell(row=r, column=2).fill = fill
            for c in range(3, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val not in (None, "", 0): ws.cell(row=r, column=c).fill = fill

    # summeringsrader (bold)
    for r in (sum_row_idx, nights_row_idx, pernight_row_idx, total_row_idx):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).font = Font(bold=True)
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

    # formel
    for col_idx in range(3, max_col + 1):
        L = get_column_letter(col_idx)
        cell = ws.cell(row=pernight_row_idx, column=col_idx)
        cell.value = f'=IFERROR({L}{sum_row_idx}/{L}{nights_row_idx},"")'
        cell.number_format = "0.0"

    # blocklinjer
    for i in range(num_species_rows):
        top_row = data_start + i * 3
        for c in range(1, max_col + 1):
            old = ws.cell(row=top_row, column=c).border
            ws.cell(row=top_row, column=c).border = Border(left=old.left, right=old.right, top=BORDER_MEDIUM, bottom=old.bottom)
    for c in range(1, max_col + 1):
        old = ws.cell(row=sum_row_idx, column=c).border
        ws.cell(row=sum_row_idx, column=c).border = Border(left=old.left, right=old.right, top=BORDER_MEDIUM, bottom=old.bottom)
        old = ws.cell(row=total_row_idx, column=c).border
        ws.cell(row=total_row_idx, column=c).border = Border(left=old.left, right=BORDER_MEDIUM, top=old.top, bottom=BORDER_MEDIUM)

    # vertikala avgränsningar
    for c in range(3, max_col + 1):
        for r in range(header_row, max_row + 1):
            old = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(left=BORDER_MEDIUM, right=old.right, top=old.top, bottom=old.bottom)
    for r in range(header_row, max_row + 1):
        oldA = ws.cell(row=r, column=1).border
        ws.cell(row=r, column=1).border = Border(left=BORDER_MEDIUM, right=oldA.right, top=oldA.top, bottom=oldA.bottom)
        oldB = ws.cell(row=r, column=2).border
        ws.cell(row=r, column=2).border = Border(left=BORDER_MEDIUM, right=oldB.right, top=oldB.top, bottom=oldB.bottom)
        oldR = ws.cell(row=r, column=max_col).border
        ws.cell(row=r, column=max_col).border = Border(left=oldR.left, right=BORDER_MEDIUM, top=oldR.top, bottom=oldR.bottom)

    wb.save(path_out); wb.close()

# Skriv båda filer
write_overview_to(out_path_nvi); format_overview(out_path_nvi, scheme_table_nvi(), num_species, file_cols)
print(f"Klar! Sparad fil (NVI): {out_path_nvi}")
write_overview_to(out_path_art); format_overview(out_path_art, scheme_table_art(), num_species, file_cols)
print(f"Klar! Sparad fil (ART): {out_path_art}")
open_file(out_path_nvi); open_file(out_path_art)

# ================== STEG 2: Diagram (linje + stapel) ==================
def _compute_ymax_for_subset(df_subset, custom_time_range):
    if df_subset.empty: return 0
    df = df_subset.copy()
    if "MANUAL ID" not in df.columns: return 0
    df["species_type_list"] = df["MANUAL ID"].map(extract_species_and_type)

    def time_to_interval(val):
        hm = _hm_from_any(val)
        if hm is None: return ""
        h, m = hm; minutes = int((m // 15) * 15)
        return f"{h:02d}:{minutes:02d}"
    time_col = detect_column(df, ["time", "tid"])
    if time_col is None: return 0
    df["interval"] = df[time_col].map(time_to_interval)

    df_long = df.explode("species_type_list")
    df_long = df_long[df_long["species_type_list"].notna()]
    if df_long.empty: return 0
    df_long[["species", "obs_type"]] = pd.DataFrame(df_long["species_type_list"].tolist(), index=df_long.index)
    df_long = df_long[df_long["species"].astype(str).str.strip().str.lower() != "noise"]
    df_long = df_long[df_long["species"].astype(str).str.strip() != ""]
    if df_long.empty: return 0

    # tidsintervall
    if custom_time_range:
        min_dt = str_to_dt(custom_time_range[0] + ":00"); max_dt = str_to_dt(custom_time_range[1] + ":00")
        if min_dt is None or max_dt is None: return 0
        min_dt = round_down_15(min_dt); max_dt = round_up_15(max_dt)
    else:
        dt_series = df[time_col].apply(str_to_dt).dropna()
        if len(dt_series) == 0:
            ints = [s for s in df["interval"].astype(str).tolist() if s and s.lower() != "nan"]
            dt_from_int = [interval_to_sortkey(s) for s in ints]
            dt_from_int = [d for d in dt_from_int if d is not None]
            if not dt_from_int: return 0
            min_dt = round_down_15(min(dt_from_int)); max_dt = round_up_15(max(dt_from_int))
        else:
            min_dt = round_down_15(min(dt_series)); max_dt = round_up_15(max(dt_series))

    # lista intervall
    all_intervals = []
    t = min_dt
    while t <= max_dt:
        all_intervals.append(t.strftime("%H:%M")); t += timedelta(minutes=15)
    all_intervals = list(dict.fromkeys(all_intervals))

    agg = df_long.groupby(["interval", "species", "obs_type"]).size().reset_index(name="antal")
    agg["interval"] = pd.Categorical(agg["interval"], categories=all_intervals, ordered=True)
    type_order = ["Socialt", "Födosökande", "Förbiflygande"]

    y_max = 0
    for sp in df_long["species"].unique():
        plot_data = (
            agg[agg["species"] == sp]
            .pivot(index="interval", columns="obs_type", values="antal")
            .fillna(0)
            .reindex(all_intervals, fill_value=0)
            .reindex(columns=type_order, fill_value=0)
        )
        if not plot_data.empty:
            y_max = max(y_max, int(plot_data.sum(axis=1).max()))
    return y_max

def compute_global_ymax_across_files(input_files_list, custom_time_range):
    y_global = 0
    for path in input_files_list:
        df = pd.read_excel(path)
        y_global = max(y_global, _compute_ymax_for_subset(df, custom_time_range))
    return max(1, math.ceil(y_global * 1.05))

def compute_global_ymax_across_files_and_nights(input_files_list, custom_time_range):
    y_global = 0
    for path in input_files_list:
        df_full = pd.read_excel(path)
        date_col = detect_column(df_full, ["date", "datum"])
        time_col = detect_column(df_full, ["time", "tid"])
        if not date_col:
            y_global = max(y_global, _compute_ymax_for_subset(df_full, custom_time_range)); continue
        df_full["__night"] = df_full.apply(lambda r: row_night_key(r[date_col], r[time_col] if time_col in r else None), axis=1)
        for night_key, sub in df_full.groupby("__night"):
            if night_key is None: continue
            y_global = max(y_global, _compute_ymax_for_subset(sub, custom_time_range))
    return max(1, math.ceil(y_global * 1.05))

def night_label_str(night_start: date) -> str:
    nxt = night_start + timedelta(days=1)
    return f"{night_start.day}/{nxt.day}.{night_start.month:02d}"

def _plot_for_subset(df_subset, custom_time_range, y_lim_global,
                     out_lines, out_stacks, colors_art, colors_nvi, type_order,
                     night_label: str | None = None):
    df = df_subset.copy()
    if "MANUAL ID" not in df.columns: return
    df["species_type_list"] = df["MANUAL ID"].map(extract_species_and_type)

    time_col = detect_column(df, ["time", "tid"])
    if time_col is None: return

    def time_to_interval(val):
        hm = _hm_from_any(val)
        if hm is None: return ""
        h, m = hm; minutes = int((m // 15) * 15)
        return f"{h:02d}:{minutes:02d}"
    df["interval"] = df[time_col].map(time_to_interval)

    df_long = df.explode("species_type_list")
    df_long = df_long[df_long["species_type_list"].notna()]
    if df_long.empty: return
    df_long[["species", "obs_type"]] = pd.DataFrame(df_long["species_type_list"].tolist(), index=df_long.index)
    df_long = df_long[df_long["species"].astype(str).str.strip().str.lower() != "noise"]
    df_long = df_long[df_long["species"].astype(str).str.strip() != ""]
    if df_long.empty: return

    # tidsintervall
    if custom_time_range:
        min_dt = str_to_dt(custom_time_range[0] + ":00"); max_dt = str_to_dt(custom_time_range[1] + ":00")
        if min_dt is None or max_dt is None: return
        min_dt = round_down_15(min_dt); max_dt = round_up_15(max_dt)
    else:
        dt_series = df[time_col].apply(str_to_dt).dropna()
        if len(dt_series) == 0:
            ints = [s for s in df["interval"].astype(str).tolist() if s and s.lower() != "nan"]
            dt_from_int = [interval_to_sortkey(s) for s in ints]
            dt_from_int = [d for d in dt_from_int if d is not None]
            if not dt_from_int: return
            min_dt = round_down_15(min(dt_from_int)); max_dt = round_up_15(max(dt_from_int))
        else:
            min_dt = round_down_15(min(dt_series)); max_dt = round_up_15(max(dt_series))

    # lista intervall
    all_intervals = []
    t = min_dt
    while t <= max_dt:
        all_intervals.append(t.strftime("%H:%M")); t += timedelta(minutes=15)
    all_intervals = list(dict.fromkeys(all_intervals))

    # Agg
    agg = df_long.groupby(["interval", "species", "obs_type"]).size().reset_index(name="antal")
    agg["interval"] = pd.Categorical(agg["interval"], categories=all_intervals, ordered=True)
    species_list = sorted(df_long["species"].unique())

    # LINJE – samlingsdiagram
    agg_line = df_long.groupby(["interval", "species"]).size().reset_index(name="antal")
    agg_line["interval"] = pd.Categorical(agg_line["interval"], categories=all_intervals, ordered=True)
    pivot_line = agg_line.pivot(index="interval", columns="species", values="antal").fillna(0)
    pivot_line = pivot_line.reindex(all_intervals, fill_value=0)
    total_obs = df_long.shape[0]

    plt.figure(figsize=(12, 6))
    ax_all = plt.gca()
    pivot_line.plot(ax=ax_all, marker='o')
    ax_all.set_xlabel("Tid (15-minutersintervall)"); ax_all.set_ylabel("Antal ljudfiler")
    title_all = "Fladdermusobservationer – alla arter"
    if night_label: title_all += f" – natt {night_label}"
    title_all += f", antal observerade beteenden: {total_obs}"
    ax_all.set_title(title_all)
    ax_all.legend(title="Art", bbox_to_anchor=(1.05, 1), loc='upper left')
    ax_all.yaxis.set_major_locator(MaxNLocator(integer=True))
    ax_all.set_xticks(range(len(pivot_line.index))); ax_all.set_xticklabels(pivot_line.index, rotation=270)
    ax_all.set_ylim(0, y_lim_global)
    plt.tight_layout(); plt.grid(True, axis='y')
    plt.savefig(os.path.join(out_lines, "alla_arter.png")); plt.close()

    # LINJE – per art
    for species in species_list:
        total_sp = int(pivot_line[species].sum())
        plt.figure(figsize=(10, 4)); ax = plt.gca()
        ax.plot(pivot_line.index, pivot_line[species], marker='o')
        ax.set_xlabel("Tid (15-minutersintervall)"); ax.set_ylabel("Antal ljudfiler")
        ax.set_title(format_title(species, total_sp, night_label))
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.set_xticks(range(len(pivot_line.index))); ax.set_xticklabels(pivot_line.index, rotation=270)
        ax.set_ylim(0, y_lim_global)
        plt.grid(True, axis='y'); plt.tight_layout()
        plt.savefig(os.path.join(out_lines, f"{safe_filename(species)}.png")); plt.close()

    # STAPEL – ART
    out_stacks_art = out_stacks + "_ART"; os.makedirs(out_stacks_art, exist_ok=True)
    for species in species_list:
        plot_data = (
            agg[agg["species"] == species]
            .pivot(index="interval", columns="obs_type", values="antal")
            .fillna(0)
            .reindex(all_intervals, fill_value=0)
            .reindex(columns=type_order, fill_value=0)
        )
        ax = plot_data.plot(kind="bar", stacked=True, color=[HEX_ART_SOC, HEX_ART_FODO, HEX_ART_FORBI], figsize=(14, 6))
        plt.xlabel("Tid (15-minutersintervall)"); plt.ylabel("Antal ljudfiler")
        plt.title(format_title(species, int(plot_data.values.sum()), night_label))
        plt.legend(title="Beteendetyper")
        ax.yaxis.set_major_locator(MaxNLocator(integer=True)); ax.set_ylim(0, y_lim_global)
        plt.xticks(rotation=270); plt.tight_layout(); plt.grid(True, axis='y')
        plt.savefig(os.path.join(out_stacks_art, f"{safe_filename(species)}.png")); plt.close()

    # STAPEL – NVI
    out_stacks_nvi = out_stacks + "_NVI"; os.makedirs(out_stacks_nvi, exist_ok=True)
    for species in species_list:
        plot_data = (
            agg[agg["species"] == species]
            .pivot(index="interval", columns="obs_type", values="antal")
            .fillna(0)
            .reindex(all_intervals, fill_value=0)
            .reindex(columns=type_order, fill_value=0)
        )
        ax = plot_data.plot(kind="bar", stacked=True, color=[HEX_NVI_SOC, HEX_NVI_FODO, HEX_NVI_FORBI], figsize=(14, 6))
        plt.xlabel("Tid (15-minutersintervall)"); plt.ylabel("Antal ljudfiler")
        plt.title(format_title(species, int(plot_data.values.sum()), night_label))
        plt.legend(title="Beteendetyper")
        ax.yaxis.set_major_locator(MaxNLocator(integer=True)); ax.set_ylim(0, y_lim_global)
        plt.xticks(rotation=270); plt.tight_layout(); plt.grid(True, axis='y')
        plt.savefig(os.path.join(out_stacks_nvi, f"{safe_filename(species)}.png")); plt.close()

def generate_summary_diagrams(input_files_list, diagrams_root, custom_time_range):
    y_lim = compute_global_ymax_across_files(input_files_list, custom_time_range)
    print(f"Global gemensam Y-max (SAMLADE): {y_lim}")
    colors_art = [HEX_ART_SOC, HEX_ART_FODO, HEX_ART_FORBI]
    colors_nvi = [HEX_NVI_SOC, HEX_NVI_FODO, HEX_NVI_FORBI]
    type_order = ["Socialt", "Födosökande", "Förbiflygande"]
    for input_file in input_files_list:
        stem = os.path.splitext(os.path.basename(input_file))[0]
        df_full = pd.read_excel(input_file)
        base_out = os.path.join(diagrams_root, f"{stem}")
        out_lines  = base_out + "_linjediagram"
        out_stacks = base_out + "_stapeldiagram"
        os.makedirs(out_lines, exist_ok=True)
        _plot_for_subset(df_full, custom_time_range, y_lim, out_lines, out_stacks,
                         colors_art, colors_nvi, type_order, night_label=None)

def generate_pernight_diagrams(input_files_list, diagrams_root, custom_time_range):
    y_lim = compute_global_ymax_across_files_and_nights(input_files_list, custom_time_range)
    print(f"Global gemensam Y-max (NATT-FÖR-NATT): {y_lim}")
    colors_art = [HEX_ART_SOC, HEX_ART_FODO, HEX_ART_FORBI]
    colors_nvi = [HEX_NVI_SOC, HEX_NVI_FODO, HEX_NVI_FORBI]
    type_order = ["Socialt", "Födosökande", "Förbiflygande"]
    for input_file in input_files_list:
        stem = os.path.splitext(os.path.basename(input_file))[0]
        df_full = pd.read_excel(input_file)
        date_col = detect_column(df_full, ["date", "datum"])
        time_col = detect_column(df_full, ["time", "tid"])
        if not date_col:
            # brak daty → potraktuj jako jedną „noc”
            base_out = os.path.join(diagrams_root, f"{stem}__natt_utan_datum")
            out_lines  = base_out + "_linjediagram"
            out_stacks = base_out + "_stapeldiagram"
            os.makedirs(out_lines, exist_ok=True)
            _plot_for_subset(df_full, custom_time_range, y_lim, out_lines, out_stacks,
                             colors_art, colors_nvi, type_order, night_label=None)
            continue
        df_full["__night"] = df_full.apply(lambda r: row_night_key(r[date_col], r[time_col] if time_col in r else None), axis=1)
        for night_start, sub in df_full.groupby("__night"):
            if night_start is None or sub.empty: continue
            night_lab = night_label_str(night_start)
            base_out = os.path.join(diagrams_root, f"{stem}__natt_{night_start.isoformat()}_{(night_start+timedelta(days=1)).isoformat()}")
            out_lines  = base_out + "_linjediagram"
            out_stacks = base_out + "_stapeldiagram"
            os.makedirs(out_lines, exist_ok=True)
            _plot_for_subset(sub, custom_time_range, y_lim, out_lines, out_stacks,
                             colors_art, colors_nvi, type_order, night_label=night_lab)

# ================== Uruchom tworzenie wykresów (jeśli wybrano) ==================
if settings["do_plots_summary"] or settings["do_plots_pernight"]:
    diagrams_base = settings["diagram_base"] or results_dir
    diagrams_root = os.path.join(diagrams_base, "diagramer")
    os.makedirs(diagrams_root, exist_ok=True)
    print(f"Resultat kommer att sparas i: {diagrams_root}")

    if settings["do_plots_summary"]:
        generate_summary_diagrams(
            input_files_list=input_files,
            diagrams_root=diagrams_root,
            custom_time_range=settings["custom_time_range"],
        )
    if settings["do_plots_pernight"]:
        generate_pernight_diagrams(
            input_files_list=input_files,
            diagrams_root=diagrams_root,
            custom_time_range=settings["custom_time_range"],
        )
