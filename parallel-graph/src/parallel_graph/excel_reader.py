"""Read timestamped MANUAL ID registrations from Excel workbooks."""

from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .models import ImportReport, Registration, SourceSpec


REQUIRED_HEADERS = {"DATE", "TIME", "MANUAL ID"}


class WorkbookFormatError(ValueError):
    """Raised when no suitable data table can be found in a workbook."""


def _normalise_header(value: Any) -> str:
    return " ".join(str(value or "").strip().upper().split())


def _find_table(workbook: Any) -> tuple[Any, int, dict[str, int]]:
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 50), values_only=True),
            start=1,
        ):
            headers = {
                _normalise_header(value): column
                for column, value in enumerate(row)
                if _normalise_header(value)
            }
            if REQUIRED_HEADERS <= headers.keys():
                return sheet, row_number, headers
    raise WorkbookFormatError(
        "Nie znaleziono arkusza z kolumnami DATE, TIME i MANUAL ID."
    )


def _date_part(value: Any, epoch: datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    raise ValueError(f"Nieprawidłowa data: {value!r}")


def _time_part(value: Any, epoch: datetime) -> time:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, timedelta):
        seconds = int(value.total_seconds()) % 86_400
        return (datetime.min + timedelta(seconds=seconds)).time()
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        if isinstance(converted, datetime):
            return converted.time()
        if isinstance(converted, time):
            return converted
    text = str(value).strip()
    for pattern in ("%H:%M:%S.%f", "%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, pattern).time()
        except ValueError:
            pass
    raise ValueError(f"Nieprawidłowy czas: {value!r}")


def combine_excel_datetime(date_value: Any, time_value: Any, epoch: datetime) -> datetime:
    """Combine Excel or textual DATE and TIME cell values."""
    return datetime.combine(_date_part(date_value, epoch), _time_part(time_value, epoch))


def split_manual_ids(value: Any) -> list[str]:
    """Expand comma-separated manual identifications into separate occurrences."""
    text = " ".join(str(value or "").strip().split())
    return [part.strip() for part in text.split(",") if part.strip()]


def read_source(source: SourceSpec) -> tuple[list[Registration], ImportReport]:
    path = Path(source.path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Nie znaleziono pliku: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet, header_row, headers = _find_table(workbook)
        date_col = headers["DATE"]
        time_col = headers["TIME"]
        species_col = headers["MANUAL ID"]
        registrations: list[Registration] = []
        examined = blank = noise = invalid_time = 0
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            date_value = row[date_col] if date_col < len(row) else None
            time_value = row[time_col] if time_col < len(row) else None
            species_value = row[species_col] if species_col < len(row) else None
            if date_value is None and time_value is None and species_value is None:
                continue
            examined += 1
            species_values = split_manual_ids(species_value)
            if not species_values:
                blank += 1
                continue
            try:
                timestamp = combine_excel_datetime(date_value, time_value, workbook.epoch)
            except (TypeError, ValueError, OverflowError):
                invalid_time += 1
                continue
            for species in species_values:
                if species.casefold() == "noise":
                    noise += 1
                    continue
                registrations.append(Registration(timestamp, source.name, species))
        report = ImportReport(
            source=source.name,
            path=path,
            sheet=sheet.title,
            rows_examined=examined,
            registrations_loaded=len(registrations),
            blank_species_skipped=blank,
            noise_skipped=noise,
            invalid_time_skipped=invalid_time,
        )
        return registrations, report
    finally:
        workbook.close()


def read_sources(sources: Iterable[SourceSpec]) -> tuple[list[Registration], list[ImportReport]]:
    registrations: list[Registration] = []
    reports: list[ImportReport] = []
    for source in sources:
        loaded, report = read_source(source)
        registrations.extend(loaded)
        reports.append(report)
    return registrations, reports
